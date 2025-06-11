"""Fixed EOL Summary sheet generator with proper risk count display."""

import pandas as pd
import os
import re
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color

from va_tool.reporting.sheets.base import BaseSheetGenerator
from va_tool.utils import get_logger


class EOLSummarySheetGenerator(BaseSheetGenerator):
    """Generator for the simplified EOL Summary sheet."""
    
    def __init__(self):
        """Initialize the generator."""
        super().__init__(title="6.0 EOL Summary")
        self.logger = get_logger()
    
    def generate(self, wb, df=None, output_dir=None, results_data=None, **kwargs):
        """Generate the EOL Summary sheet."""
        self.logger.info("Generating EOL Summary sheet")
        ws = super().generate(wb)
        
        # Add title
        self.add_title(
            ws, "End of Life Components Summary", 
            font_size=16, bold=True, merge_range='A1:H1'
        )
        
        # Extract SEoL data
        if df is not None:
            # Filter for SEoL entries
            seol_df = df[df["Name"].str.contains("SEoL", case=False, na=False)]
            
            if not seol_df.empty:
                self.logger.info(f"Found {len(seol_df)} SEoL entries")
                
                # Process the data
                summary_data = self.process_eol_data(seol_df)
                
                # Write the summary sections
                self.write_overview_metrics(ws, summary_data)
                
                # Create visualization charts
                self.create_summary_charts(ws, summary_data)
            else:
                ws['A3'] = "No SEoL components found in the vulnerability data."
        else:
            ws['A3'] = "No vulnerability data provided."
        
        return ws
    
    def process_eol_data(self, seol_df):
        """Process EOL data to create a simplified high-level summary."""
        # 1. Unique components vs. total instances
        unique_components = seol_df.drop_duplicates(subset=['Plugin ID', 'Name'])
        total_components = len(unique_components)
        total_instances = len(seol_df)
        
        # 2. Affected hosts
        affected_hosts = seol_df['Host'].nunique()
        
        # 3. Extract software types and versions
        software_versions = []
        for _, row in seol_df.iterrows():
            name = str(row['Name']) if pd.notna(row.get('Name')) else ""
            plugin_id = str(row.get('Plugin ID', ''))
            host = str(row.get('Host', ''))
            risk = str(row.get('Risk', ''))
            
            # Extract software type and version
            software_type, version = self.extract_software_and_version(name)
            
            software_versions.append({
                'Software': software_type,
                'Version': version,
                'Plugin ID': plugin_id,
                'Host': host, 
                'Risk': risk,
                'Name': name
            })
        
        # Convert to DataFrame
        versions_df = pd.DataFrame(software_versions)
        
        # 4. Count software types and versions
        software_types = versions_df['Software'].nunique()
        unique_versions = len(versions_df.groupby(['Software', 'Version']))
        
        # 5. Analyze by risk level
        risk_counts = versions_df['Risk'].value_counts().reset_index()
        risk_counts.columns = ['Risk', 'Count']
        
        # Ensure we have all risk levels even if count is 0
        all_risks = ["Critical", "High", "Medium", "Low"]
        risk_dict = {row['Risk']: row['Count'] for _, row in risk_counts.iterrows()}
        
        # Create a complete risk_counts DataFrame with all risk levels
        complete_risk_counts = []
        for risk in all_risks:
            complete_risk_counts.append({'Risk': risk, 'Count': risk_dict.get(risk, 0)})
        
        risk_counts_df = pd.DataFrame(complete_risk_counts)
        
        # 6. Get all software types with counts
        software_counts = versions_df['Software'].value_counts().reset_index()
        software_counts.columns = ['Software', 'Count']
        
        # Return all processed data
        return {
            'total_components': total_components,
            'total_instances': total_instances,
            'affected_hosts': affected_hosts,
            'software_types': software_types,
            'unique_versions': unique_versions,
            'risk_counts': risk_counts_df,
            'software_counts': software_counts  # Include all software types (not just top 10)
        }
    
    def extract_software_and_version(self, name):
        """Extract software type and version from a SEoL name."""
        # Default values
        software_type = "Unknown"
        version = "Unknown"
        
        # Common software patterns
        software_patterns = [
            # Unsupported Version of X
            (r'Unsupported\s+Version\s+of\s+([\w\.\s\-\(\)]+)', 1),
            # X Unsupported Version
            (r'([\w\.\s\-\(\)]+)\s+Unsupported\s+Version', 1),
            # X SEoL
            (r'([\w\.\s\-\(\)]+)\s+SEoL', 1),
            # SEoL X
            (r'SEoL\s+([\w\.\s\-\(\)]+)', 1),
        ]
        
        # Try each pattern to extract software type
        for pattern, group in software_patterns:
            match = re.search(pattern, name, re.IGNORECASE)
            if match:
                software_type = match.group(group).strip()
                break
        
        # If no match, use the first word after "Unsupported" or "SEoL"
        if software_type == "Unknown":
            clean_name = re.sub(r'^(Unsupported|SEoL)\s+', '', name)
            words = clean_name.split()
            if words:
                software_type = words[0]
        
        # Version patterns
        version_patterns = [
            r'(\d+\.\d+\.\d+(\.\d+)?)',  # Common format x.y.z.w
            r'version\s+(\d+\.\d+(\.\d+)?(\.\d+)?)',  # "version X.Y.Z"
            r'v(\d+\.\d+(\.\d+)?(\.\d+)?)',  # "vX.Y.Z"
            r'(\d+\.\d+)\s',  # "X.Y " (with space after)
            r'(\d+\.\d+[a-z]?)',  # Format x.y or x.ya
            r'Windows\s+Server\s+(\d{4})',  # Windows Server 2012, 2016, etc.
            r'Windows\s+(\d+)',  # Windows 7, 8, 10, 11, etc.
        ]
        
        # Try each pattern to extract version
        for pattern in version_patterns:
            match = re.search(pattern, name, re.IGNORECASE)
            if match:
                version = match.group(1)
                break
        
        return software_type, version
    
    def write_overview_metrics(self, ws, summary_data):
        """Write the key metrics at the top of the page in a clean, visual format."""
        # Add section title
        ws['A3'] = "EOL Summary Statistics"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A3'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A3'].font = Font(bold=True, size=12, color="FFFFFF")
        ws.merge_cells('A3:H3')
        
        # Create a visually appealing metrics row
        metrics = [
            {"name": "Total Unique EOL\nComponents", "value": summary_data['total_components']},
            {"name": "Software Types\nAffected", "value": summary_data['software_types']},
            {"name": "Hosts with EOL\nComponents", "value": summary_data['affected_hosts']},
            {"name": "Unique EOL\nVersions", "value": summary_data['unique_versions']}
        ]
        
        # Create a visually appealing metrics display
        for i, metric in enumerate(metrics, 1):
            # Column for each metric
            col = i * 2 - 1  # 1, 3, 5, 7
            
            # Metric name
            name_cell = ws.cell(row=4, column=col, value=metric["name"])
            name_cell.font = Font(bold=True)
            name_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            name_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Add borders
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            name_cell.border = border
            
            ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
            
            # Metric value
            value_cell = ws.cell(row=5, column=col, value=metric["value"])
            value_cell.font = Font(bold=True, size=16)
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            value_cell.border = border
            
            # Special highlight for first metric
            if i == 1:
                value_cell.font = Font(bold=True, size=16, color="9C0006")
                value_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
        
        # Set row heights
        ws.row_dimensions[4].height = 40
        ws.row_dimensions[5].height = 40
        
        # Add risk distribution summary
        risk_title_row = 7
        ws.cell(row=risk_title_row, column=1, value="Risk Distribution")
        ws.cell(row=risk_title_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=risk_title_row, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.cell(row=risk_title_row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
        ws.merge_cells(f'A{risk_title_row}:H{risk_title_row}')
        
        # Risk levels, colors, and fonts
        risk_levels = ["Critical", "High", "Medium", "Low"]
        risk_colors = ["FFC7CE", "FFEB9C", "FFFFCC", "C6EFCE"]
        risk_fonts = ["9C0006", "9C5700", "9C8000", "006100"]
        
        # Create a common border style
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Get risk counts - convert to a dictionary for easy access
        risk_data = {row['Risk']: row['Count'] for _, row in summary_data['risk_counts'].iterrows()}
        total_risk_count = sum(risk_data.values())
        
        # Risk distribution header row
        header_row = risk_title_row + 1
        for i, risk in enumerate(risk_levels, 1):
            col = i * 2 - 1  # 1, 3, 5, 7
            
            # Risk level header
            header_cell = ws.cell(row=header_row, column=col, value=risk)
            header_cell.font = Font(bold=True, color=risk_fonts[i-1])
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            header_cell.fill = PatternFill(start_color=risk_colors[i-1], end_color=risk_colors[i-1], fill_type="solid")
            header_cell.border = border
            
            ws.merge_cells(start_row=header_row, start_column=col, end_row=header_row, end_column=col+1)
        
        # Risk distribution count row
        count_row = header_row + 1
        for i, risk in enumerate(risk_levels, 1):
            col = i * 2 - 1  # 1, 3, 5, 7
            
            # Risk count - ensure it exists in the data or default to 0
            count = risk_data.get(risk, 0)
            count_cell = ws.cell(row=count_row, column=col, value=count)
            count_cell.font = Font(bold=True, size=14)
            count_cell.alignment = Alignment(horizontal="center", vertical="center")
            count_cell.border = border
            
            ws.merge_cells(start_row=count_row, start_column=col, end_row=count_row, end_column=col+1)
        
        # Risk distribution percentage row
        percent_row = count_row + 1
        for i, risk in enumerate(risk_levels, 1):
            col = i * 2 - 1  # 1, 3, 5, 7
            
            # Risk percentage
            count = risk_data.get(risk, 0)
            percentage = f"{(count / total_risk_count * 100):.1f}%" if total_risk_count > 0 else "0%"
            
            percent_cell = ws.cell(row=percent_row, column=col, value=percentage)
            percent_cell.font = Font(italic=True)
            percent_cell.alignment = Alignment(horizontal="center", vertical="center")
            percent_cell.border = border
            
            ws.merge_cells(start_row=percent_row, start_column=col, end_row=percent_row, end_column=col+1)
        
        # Add total row
        total_row = percent_row + 1
        
        # Total label
        total_label = ws.cell(row=total_row, column=1, value="Total")
        total_label.font = Font(bold=True)
        total_label.alignment = Alignment(horizontal="center", vertical="center")
        total_label.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_label.border = border
        
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        
        # Total count
        total_count = ws.cell(row=total_row, column=5, value=total_risk_count)
        total_count.font = Font(bold=True, size=14)
        total_count.alignment = Alignment(horizontal="center", vertical="center")
        total_count.border = border
        
        ws.merge_cells(start_row=total_row, start_column=5, end_row=total_row, end_column=8)
        
        # Set row heights for risk distribution
        ws.row_dimensions[header_row].height = 25
        ws.row_dimensions[count_row].height = 30
        ws.row_dimensions[percent_row].height = 25
        ws.row_dimensions[total_row].height = 25
        
        # Log the risk data for debugging
        self.logger.info(f"Risk data: {risk_data}")
        self.logger.info(f"Total risk count: {total_risk_count}")
    
    def create_summary_charts(self, ws, summary_data):
        """Create visualization charts for the summary data."""
        # Determine the starting row for charts (after risk distribution)
        chart_start_row = 13  # This is after the risk distribution section
        
        # Add chart titles
        ws.cell(row=chart_start_row, column=1, value="EOL Components by Risk Level")
        ws.cell(row=chart_start_row, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(f'A{chart_start_row}:D{chart_start_row}')
        
        ws.cell(row=chart_start_row, column=5, value="Top EOL Software Types")
        ws.cell(row=chart_start_row, column=5).font = Font(bold=True, size=12)
        ws.merge_cells(f'E{chart_start_row}:H{chart_start_row}')
        
        # 1. Risk Distribution Pie Chart
        self.create_risk_pie_chart(ws, summary_data['risk_counts'], chart_start_row + 1)
        
        # 2. Software Distribution Chart (updated to show all software)
        self.create_software_bar_chart(ws, summary_data['software_counts'], chart_start_row + 1)
    
    def create_risk_pie_chart(self, ws, risk_counts, start_row):
        """Create a pie chart showing risk distribution."""
        # Only include risks with count > 0 for the chart
        chart_data = risk_counts[risk_counts['Count'] > 0]
        
        # If no data with count > 0, just use all data
        if len(chart_data) == 0:
            chart_data = risk_counts
        
        # Prepare data for the chart
        chart_data_row = 60
        ws.cell(row=chart_data_row, column=1, value="Risk")
        ws.cell(row=chart_data_row, column=2, value="Count")
        
        for i, (_, row) in enumerate(chart_data.iterrows(), 1):
            ws.cell(row=chart_data_row+i, column=1, value=row['Risk'])
            ws.cell(row=chart_data_row+i, column=2, value=row['Count'])
        
        # Create pie chart
        pie = PieChart()
        pie.title = None  # Already have a title in the worksheet
        
        # Define data range for the chart
        labels = Reference(ws, min_col=1, min_row=chart_data_row+1, max_row=chart_data_row+len(chart_data))
        data = Reference(ws, min_col=2, min_row=chart_data_row, max_row=chart_data_row+len(chart_data))
        
        # Add data to the chart
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        
        # Add data labels showing percentages
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        
        # Set chart size and position
        pie.width = 15
        pie.height = 10
        
        # Add chart to worksheet
        ws.add_chart(pie, f"A{start_row}")
    
    def create_software_bar_chart(self, ws, software_counts, start_row):
        """Create a bar chart showing top software types."""
        # Take top 10 for cleaner chart (instead of top 5)
        top_software = software_counts.head(10)
        
        # Prepare data for the chart
        chart_data_row = 75
        ws.cell(row=chart_data_row, column=4, value="Software")
        ws.cell(row=chart_data_row, column=5, value="Count")
        
        for i, (_, row) in enumerate(top_software.iterrows(), 1):
            ws.cell(row=chart_data_row+i, column=4, value=row['Software'])
            ws.cell(row=chart_data_row+i, column=5, value=row['Count'])
        
        # Create bar chart
        chart = BarChart()
        chart.title = None  # Already have a title in the worksheet
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Software Type"
        
        # Define data range for the chart
        data = Reference(ws, min_col=5, min_row=chart_data_row, max_row=chart_data_row+len(top_software))
        cats = Reference(ws, min_col=4, min_row=chart_data_row+1, max_row=chart_data_row+len(top_software))
        
        # Add data to chart
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Add data labels
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        
        # Set chart size and position - increase size to avoid overlapping
        chart.width = 20  # Increased from 15
        chart.height = 14  # Increased from 10
        
        # Add chart to worksheet
        ws.add_chart(chart, f"E{start_row}")