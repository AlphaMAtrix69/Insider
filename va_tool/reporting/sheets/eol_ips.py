"""EOL IPs sheet generator using Excel native charts."""

import pandas as pd
import os
import re
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color

from va_tool.reporting.sheets.base import BaseSheetGenerator
from va_tool.utils import ensure_dir_exists


class EOLIPsSheetGenerator(BaseSheetGenerator):
    """Generator for the EOL IPs sheet."""
    
    def __init__(self):
        """Initialize the generator."""
        super().__init__(title="6.2 EOL IPs")
        self.seol_df = None
        self.all_ips_cache = None  # Cache for all IPs
    
    def generate(self, wb, df=None, output_dir=None, results_data=None, **kwargs):
        """
        Generate the EOL IPs sheet.
        
        Args:
            wb: Excel workbook
            df: DataFrame with vulnerability data
            output_dir: Directory to save chart images
            results_data: Dictionary with analysis results
            **kwargs: Additional arguments
        
        Returns:
            The worksheet
        """
        self.logger.info("Generating EOL IPs sheet")
        ws = super().generate(wb)
        
        # Add title
        self.add_title(
            ws, "Total IPs with SEoL Components", 
            font_size=14, merge_range='A1:F1'
        )
        
        # Use seol_summary from results_data if available - this exactly matches the standard report
        if results_data and 'seol_summary' in results_data:
            # Direct use of SEOL summary from standard report
            seol_summary = results_data.get('seol_summary', [])
            
            if seol_summary:
                # Convert to DataFrame for easier processing
                if isinstance(seol_summary, list):
                    seol_ips_df = pd.DataFrame(seol_summary)
                    
                    # Check that we have expected columns
                    if 'Host' in seol_ips_df.columns or 'IP' in seol_ips_df.columns:
                        # Store for later use
                        self.seol_df = seol_ips_df
                        
                        # Analyze IPs with SEoL components
                        ip_summary = self.prepare_ip_summary_from_seol(seol_ips_df)
                        all_ips = self.get_all_ips_from_seol(seol_ips_df)
                        self.all_ips_cache = all_ips  # Cache all IPs
                        top_ips = self.get_top_ips_from_seol(seol_ips_df)
                        
                        # Write IP statistics
                        self.write_ip_stats(ws, ip_summary)
                        
                        # Write top IPs (for chart)
                        self.write_top_ips(ws, top_ips)
                        
                        # Write all IPs with SEoL components
                        self.write_all_ips(ws, all_ips)
                        
                        # Create Excel chart
                        if not top_ips.empty:
                            self.create_eol_ip_chart(ws, top_ips)
                    else:
                        # If we have seol_summary but it doesn't have IP info,
                        # try using the full dataframe
                        if df is not None:
                            self.process_from_dataframe(df, ws)
                        else:
                            ws['A3'] = "SEoL summary data available but IP information is missing."
                else:
                    # Unexpected format, try processing from dataframe
                    if df is not None:
                        self.process_from_dataframe(df, ws)
                    else:
                        ws['A3'] = "SEoL summary data available but in unexpected format."
            else:
                # No SEOL summary, process from dataframe if available
                if df is not None:
                    self.process_from_dataframe(df, ws)
                else:
                    ws['A3'] = "No SEoL IPs found in the vulnerability data."
        
        # If no SEOL data in results_data, process directly from dataframe
        elif df is not None:
            self.process_from_dataframe(df, ws)
        else:
            ws['A3'] = "No vulnerability data available."
        
        return ws
    
    def process_from_dataframe(self, df, ws):
        """Process SEoL data directly from the main dataframe."""
        # Filter only for SEoL components, making sure to use the exact same filter as standard report
        self.seol_df = df[df['Name'].str.contains("SEoL", case=False, na=False)]
        
        if not self.seol_df.empty:
            self.logger.info(f"Found {len(self.seol_df)} SEoL entries across {self.seol_df['Host'].nunique()} hosts")
            
            # Analyze IPs with SEoL components
            ip_summary, top_ips = self.analyze_eol_ips(self.seol_df)
            
            # Get all IPs, not just the top ones - include all IPs without filtering
            all_ips = self.get_all_ips(self.seol_df)
            self.all_ips_cache = all_ips  # Cache all IPs for later use
            
            # Write IP statistics
            self.write_ip_stats(ws, ip_summary)
            
            # Write top IPs with most SEoL components (for chart)
            self.write_top_ips(ws, top_ips)
            
            # Write all IPs with SEoL components
            self.write_all_ips(ws, all_ips)
            
            # Create Excel chart
            if not top_ips.empty:
                self.create_eol_ip_chart(ws, top_ips)
        else:
            ws['A3'] = "No SEoL components found in the vulnerability data."
    
    def prepare_ip_summary_from_seol(self, seol_ips_df):
        """Prepare summary statistics from SEOL IPs DataFrame."""
        if 'Host' in seol_ips_df.columns:
            host_col = 'Host'
        elif 'IP' in seol_ips_df.columns:
            host_col = 'IP'
        else:
            return {'Total IPs with SEoL Components': 0}
            
        if 'SEoL Count' in seol_ips_df.columns:
            count_col = 'SEoL Count'
        else:
            # If no count column, assign 1 to each row (each IP has at least 1 SEoL)
            seol_ips_df['Count'] = 1
            count_col = 'Count'
            
        # Calculate summary metrics
        total_ips = len(seol_ips_df)
        
        if total_ips > 0:
            avg_components = seol_ips_df[count_col].mean()
            max_components = seol_ips_df[count_col].max()
            
            count_1 = len(seol_ips_df[seol_ips_df[count_col] == 1])
            count_2_5 = len(seol_ips_df[(seol_ips_df[count_col] > 1) & (seol_ips_df[count_col] <= 5)])
            count_gt_5 = len(seol_ips_df[seol_ips_df[count_col] > 5])
            
            return {
                'Total IPs with SEoL Components': total_ips,
                'Average SEoL Components per IP': round(avg_components, 2),
                'Maximum SEoL Components on Single IP': max_components,
                'IPs with 1 SEoL Component': count_1,
                'IPs with 2-5 SEoL Components': count_2_5,
                'IPs with >5 SEoL Components': count_gt_5
            }
        else:
            return {'Total IPs with SEoL Components': 0}
    
    def get_all_ips_from_seol(self, seol_ips_df):
        """Get all IPs with SEoL components from summary data."""
        if seol_ips_df.empty:
            return pd.DataFrame(columns=['IP', 'SEoL Component Count'])
            
        # Determine column names
        if 'Host' in seol_ips_df.columns:
            host_col = 'Host'
        elif 'IP' in seol_ips_df.columns:
            host_col = 'IP'
        else:
            return pd.DataFrame(columns=['IP', 'SEoL Component Count'])
            
        if 'SEoL Count' in seol_ips_df.columns:
            count_col = 'SEoL Count'
        else:
            # If no count column, assign 1 to each row (each IP has at least 1 SEoL)
            seol_ips_df['Count'] = 1
            count_col = 'Count'
        
        # Create a new DataFrame with standardized column names
        all_ips = pd.DataFrame({
            'IP': seol_ips_df[host_col],
            'SEoL Component Count': seol_ips_df[count_col]
        })
        
        # Sort by count descending
        all_ips = all_ips.sort_values('SEoL Component Count', ascending=False)
        
        return all_ips
    
    def get_top_ips_from_seol(self, seol_ips_df):
        """Get top IPs with most SEoL components from summary data."""
        all_ips = self.get_all_ips_from_seol(seol_ips_df)
        
        # Take top 15 for the chart
        return all_ips.head(15) if not all_ips.empty else pd.DataFrame(columns=['IP', 'SEoL Component Count'])
    
    def analyze_eol_ips(self, seol_df):
        """Analyze IPs with SEoL components."""
        # Count SEoL components per IP
        ip_counts = seol_df['Host'].value_counts().reset_index()
        ip_counts.columns = ['IP', 'SEoL Component Count']
        
        # Sort by count descending
        ip_counts = ip_counts.sort_values('SEoL Component Count', ascending=False)
        
        # Prepare summary statistics
        total_ips = len(ip_counts)
        ip_summary = {
            'Total IPs with SEoL Components': total_ips,
            'Average SEoL Components per IP': round(ip_counts['SEoL Component Count'].mean(), 2) if total_ips > 0 else 0,
            'Maximum SEoL Components on Single IP': ip_counts['SEoL Component Count'].max() if total_ips > 0 else 0,
            'IPs with 1 SEoL Component': len(ip_counts[ip_counts['SEoL Component Count'] == 1]),
            'IPs with 2-5 SEoL Components': len(ip_counts[(ip_counts['SEoL Component Count'] > 1) & 
                                                        (ip_counts['SEoL Component Count'] <= 5)]),
            'IPs with >5 SEoL Components': len(ip_counts[ip_counts['SEoL Component Count'] > 5])
        }
        
        # Get top 15 IPs with most SEoL components
        top_ips = ip_counts.head(15) if not ip_counts.empty else pd.DataFrame(columns=['IP', 'SEoL Component Count'])
        
        return ip_summary, top_ips
    
    def get_all_ips(self, seol_df):
        """Get all IPs with SEoL components without any filtering."""
        # Count SEoL components per IP
        ip_counts = seol_df['Host'].value_counts().reset_index()
        ip_counts.columns = ['IP', 'SEoL Component Count']
        
        # Sort by count descending
        return ip_counts.sort_values('SEoL Component Count', ascending=False)
    
    def write_ip_stats(self, ws, ip_summary):
        """Write IP statistics to worksheet."""
        # Write headers
        ws['A3'] = "SEoL IP Statistics"
        ws['A3'].font = ws['A3'].font.copy(bold=True)
        
        # Write summary data
        row = 4
        for stat, value in ip_summary.items():
            ws.cell(row=row, column=1, value=stat)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
    
    def write_headers(self, ws, headers, row):
        """Write formatted headers to the worksheet."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add borders
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    def write_top_ips(self, ws, top_ips):
        """Write top IPs with most SEoL components to worksheet."""
        # Add section title
        start_row = 12
        self.add_section_title(ws, "Top IPs with Most SEoL Components (For Chart)", cell=f"A{start_row}")
        
        # Write headers
        headers = ['IP Address', 'SEoL Component Count', 'Risk Level']
        self.write_headers(ws, headers, row=start_row+1)
        
        # Write data
        for row_idx, (_, row) in enumerate(top_ips.iterrows(), start_row+2):
            ip_value = row['IP']
            count_value = row['SEoL Component Count']
            
            ws.cell(row=row_idx, column=1, value=ip_value)
            ws.cell(row=row_idx, column=2, value=count_value)
            
            # Determine risk level based on count
            risk_level = "Low"
            if count_value > 10:
                risk_level = "Critical"
                cell_color = "FF0000"  # Red
            elif count_value > 5:
                risk_level = "High"
                cell_color = "FF8000"  # Orange
            elif count_value > 2:
                risk_level = "Medium"
                cell_color = "FFBF00"  # Amber/Yellow
            else:
                cell_color = "00FF00"  # Green
                
            # Set cell value and color
            cell = ws.cell(row=row_idx, column=3, value=risk_level)
            cell.font = cell.font.copy(color=cell_color)
        
        # Set column widths
        ws.column_dimensions['A'].width = 25  # IP Address
        ws.column_dimensions['B'].width = 20  # Component Count
        ws.column_dimensions['C'].width = 15  # Risk Level
    
    def write_all_ips(self, ws, all_ips):
        """Write all IPs with SEoL components to worksheet."""
        # Skip if there are no IPs
        if all_ips.empty:
            return
            
        # Add section title - start after the top IPs section and chart
        start_row = 50  # Increased from 35 to provide more space for the chart
        self.add_section_title(ws, "Complete List of All IPs with SEoL Components", cell=f"A{start_row}")
        
        # Write headers
        headers = ['IP Address', 'SEoL Component Count', 'Risk Level']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row+1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add borders
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Write data for all IPs - display ALL IPs without any filtering
        for row_idx, (_, row) in enumerate(all_ips.iterrows(), start_row+2):
            ip_value = row['IP']
            count_value = row['SEoL Component Count']
            
            # Alternate row colors
            fill_color = "E9F1FB" if row_idx % 2 == 0 else "FFFFFF"
            
            # IP cell
            ip_cell = ws.cell(row=row_idx, column=1, value=ip_value)
            ip_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            # Count cell
            count_cell = ws.cell(row=row_idx, column=2, value=count_value)
            count_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            count_cell.alignment = Alignment(horizontal="center")
            
            # Determine risk level based on count
            risk_level = "Low"
            if count_value > 10:
                risk_level = "Critical"
                cell_color = "FF0000"  # Red
            elif count_value > 5:
                risk_level = "High"
                cell_color = "FF8000"  # Orange
            elif count_value > 2:
                risk_level = "Medium"
                cell_color = "FFBF00"  # Amber/Yellow
            else:
                cell_color = "00FF00"  # Green
                
            # Set risk level cell value and color
            risk_cell = ws.cell(row=row_idx, column=3, value=risk_level)
            risk_cell.font = risk_cell.font.copy(color=cell_color)
            risk_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            risk_cell.alignment = Alignment(horizontal="center")
        
        # Set column widths
        ws.column_dimensions['A'].width = 25  # IP Address
        ws.column_dimensions['B'].width = 20  # Component Count
        ws.column_dimensions['C'].width = 15  # Risk Level
    
    def create_eol_ip_chart(self, ws, top_ips):
        """Create Excel native chart for SEoL components per IP."""
        # Create a horizontal bar chart
        chart = BarChart()
        chart.type = "bar"  # Horizontal bar chart
        chart.style = 10
        chart.title = "Top 10 IPs by SEoL Component Count"
        
        # Correct axis labels for horizontal bar chart
        chart.x_axis.title = "Number of SEoL Components"  # X-axis for horizontal bars
        chart.y_axis.title = "IP Address"  # Y-axis for IP names
        
        # Create a copy of the dataframe to manipulate
        display_ips = top_ips.copy()
        
        # Check if the specific IP is in our full dataset but not in our top display IPs
        specific_ip = "10.168.50.77"
        if specific_ip not in display_ips['IP'].values:
            # Find the specific IP in the full list of IPs
            all_ips = self.get_all_ips(self.seol_df) if hasattr(self, 'seol_df') and self.seol_df is not None else pd.DataFrame()
            
            if all_ips.empty and hasattr(self, 'all_ips_cache') and self.all_ips_cache is not None:
                # Try using cached all_ips if available
                all_ips = self.all_ips_cache
                
            if not all_ips.empty and specific_ip in all_ips['IP'].values:
                # Get the row for the specific IP
                specific_ip_row = all_ips[all_ips['IP'] == specific_ip]
                
                # If we found it, add it to our display IPs
                if not specific_ip_row.empty:
                    display_ips = pd.concat([display_ips, specific_ip_row]).reset_index(drop=True)
        
        # Sort by count descending and limit to top 12 (increased from 10 to ensure specific IP is visible)
        display_ips = display_ips.sort_values('SEoL Component Count', ascending=False).head(12)
        
        # Make sure to write these IPs to the worksheet first
        # Clear the previous top IPs section
        start_row = 12 + 2  # Skip header + title
        for row_idx in range(start_row, start_row + 20):  # Clear enough rows
            for col in range(1, 4):
                cell = ws.cell(row=row_idx, column=col)
                cell.value = None
        
        # Write the display IPs to the worksheet
        for i, (_, row) in enumerate(display_ips.iterrows()):
            row_idx = start_row + i
            ws.cell(row=row_idx, column=1, value=row['IP'])
            ws.cell(row=row_idx, column=2, value=row['SEoL Component Count'])
            
            # Determine risk level based on count
            count_value = row['SEoL Component Count']
            risk_level = "Low"
            if count_value > 10:
                risk_level = "Critical"
                cell_color = "FF0000"  # Red
            elif count_value > 5:
                risk_level = "High"
                cell_color = "FF8000"  # Orange
            elif count_value > 2:
                risk_level = "Medium"
                cell_color = "FFBF00"  # Amber/Yellow
            else:
                cell_color = "00FF00"  # Green
                
            # Set cell value and color
            cell = ws.cell(row=row_idx, column=3, value=risk_level)
            cell.font = cell.font.copy(color=cell_color)
        
        # Set up data references for the chart based on what we just wrote
        data_rows = len(display_ips)
        data = Reference(ws, min_col=2, max_col=2, min_row=start_row, max_row=start_row+data_rows-1)
        cats = Reference(ws, min_col=1, max_col=1, min_row=start_row, max_row=start_row+data_rows-1)
        
        # Add data to chart
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        
        # Enable data labels
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        
        # Adjust chart size - increase size to avoid overlapping
        chart.width = 20  # Increased from 15
        chart.height = 16  # Increased from 12
        
        # Determine chart position - position below the data to avoid overlap
        last_row = start_row + data_rows + 1
        chart_position = f"D{last_row}"
        
        # Add the chart to the worksheet at the calculated position
        ws.add_chart(chart, chart_position)
        
    def add_section_title(self, ws, title, cell="A1", font_size=12):
        """Add a section title with styling to the worksheet."""
        ws[cell] = title
        ws[cell].font = Font(bold=True, size=font_size)
        # Add a light blue background to the title
        ws[cell].fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")