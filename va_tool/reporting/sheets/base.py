"""Base sheet generator that other sheet generators can inherit from."""

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from va_tool.utils import style_header_cell, set_column_widths, get_logger

logger = get_logger()


class BaseSheetGenerator:
    """Base class for Excel sheet generators."""
    
    def __init__(self, title=None):
        """
        Initialize the sheet generator.
        
        Args:
            title: Title for the sheet
        """
        self.title = title
        self.logger = get_logger()
        
        # KPMG color scheme
        self.kpmg_blue = "00338D"
        self.kpmg_light_blue = "DCE6F1"
        self.kpmg_dark_blue = "005EB8"
        self.kpmg_medium_blue = "0091DA"
        self.kpmg_accent_blue = "00A3E0"
        self.kpmg_light_accent = "BDD6E6"
        
        # Pastel risk colors for better readability
        self.risk_colors = {
            "Critical": "FF7D7D",  # Darker pastel red
            "High": "FFB366",      # Darker pastel orange
            "Medium": "FFDA66",    # Darker pastel yellow
            "Low": "99CC99"        # Darker pastel green
        }
    
    def generate(self, wb, df=None, **kwargs):
        """
        Generate the Excel sheet.
        
        Args:
            wb: Excel workbook
            df: DataFrame with data (optional)
            **kwargs: Additional arguments
        
        Returns:
            The worksheet
        """
        self.logger.info(f"Generating sheet: {self.title or 'Untitled'}")
        
        # Create the sheet
        ws = wb.create_sheet(title=self.title)
        
        # Set up the sheet with KPMG styling
        self.setup_sheet(ws, **kwargs)
        
        # Generate content if DataFrame is provided
        if df is not None:
            self.generate_content(ws, df, **kwargs)
        
        return ws
    
    def setup_sheet(self, ws, **kwargs):
        """
        Set up the worksheet with basic formatting.
        
        Args:
            ws: The worksheet
            **kwargs: Additional arguments
        """
        # Set column widths (default: 20 columns, width 15)
        set_column_widths(ws, kwargs.get('num_columns', 20), kwargs.get('column_width', 15))
        
        # Add default KPMG styling to the sheet
        self.apply_kpmg_theme(ws)
    
    def apply_kpmg_theme(self, ws):
        """
        Apply KPMG theme to the worksheet.
        
        Args:
            ws: The worksheet
        """
        # Set sheet properties for better printing
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        # Set default font for the entire sheet
        for row in range(1, 1000):  # Apply to a large number of rows
            for col in range(1, 30):  # Apply to a large number of columns
                cell = ws.cell(row=row, column=col)
                if not cell.font:
                    cell.font = Font(name="Calibri", size=10)
    
    def generate_content(self, ws, df, **kwargs):
        """
        Generate content for the worksheet.
        
        Args:
            ws: The worksheet
            df: DataFrame with data
            **kwargs: Additional arguments
        """
        # This method should be implemented by subclasses
        pass
    
    def add_title(self, ws, title, cell="A1", font_size=14, bold=True, merge_range=None):
        """
        Add a title to the worksheet with KPMG styling.
        
        Args:
            ws: The worksheet
            title: Title text
            cell: Cell reference for the title
            font_size: Font size
            bold: Whether the title should be bold
            merge_range: Range to merge (e.g., "A1:H1")
        """
        ws[cell] = title
        
        # Apply KPMG styling
        ws[cell].font = Font(size=font_size, bold=bold, color="FFFFFF")
        ws[cell].fill = PatternFill(start_color=self.kpmg_blue, end_color=self.kpmg_blue, fill_type="solid")
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")
        
        if merge_range:
            ws.merge_cells(merge_range)
            
        # Set row height for title
        cell_coord = coordinate_from_string(cell)
        row_num = cell_coord[1]
        ws.row_dimensions[row_num].height = 30
    
    def add_section_title(self, ws, title, cell="A3", bold=True, with_border=True):
        """
        Add a section title to the worksheet with KPMG styling.
        
        Args:
            ws: The worksheet
            title: Title text
            cell: Cell reference for the title
            bold: Whether the title should be bold
            with_border: Whether to add a border
        """
        ws[cell] = title
        ws[cell].font = Font(bold=bold, color=self.kpmg_dark_blue)
        ws[cell].fill = PatternFill(start_color=self.kpmg_light_blue, end_color=self.kpmg_light_blue, fill_type="solid")
        
        if with_border:
            ws[cell].border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    def write_headers(self, ws, headers, row=3, start_col=1):
        """
        Write headers to the worksheet with KPMG styling.
        
        Args:
            ws: The worksheet
            headers: List of header texts
            row: Row number
            start_col: Starting column number
        """
        for col_idx, header in enumerate(headers, start_col):
            cell = ws.cell(row=row, column=col_idx, value=header)
            
            # Apply KPMG styling
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.kpmg_dark_blue, end_color=self.kpmg_dark_blue, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Add border
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Adjust column width based on header length
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(10, min(30, len(str(header)) + 2))
    
    def write_data_rows(self, ws, data, row_start=4, col_start=1, apply_styles=True):
        """
        Write data rows to the worksheet with optional styling.
        
        Args:
            ws: The worksheet
            data: List of lists or DataFrame
            row_start: Starting row number
            col_start: Starting column number
            apply_styles: Whether to apply KPMG styling
        """
        # If data is a DataFrame, convert it to a list of lists
        if hasattr(data, 'values'):
            data = data.values.tolist()
        
        for row_idx, row_data in enumerate(data, row_start):
            for col_idx, value in enumerate(row_data, col_start):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                if apply_styles:
                    # Add border
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Add alternating row colors
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color=self.kpmg_light_accent, end_color=self.kpmg_light_accent, fill_type="solid")
    
    def apply_conditional_formatting(self, ws, data, col_name_to_idx, start_row, end_row):
        """
        Apply conditional formatting to data cells based on content.
        
        Args:
            ws: The worksheet
            data: DataFrame with data
            col_name_to_idx: Dictionary mapping column names to column indices
            start_row: Starting row number
            end_row: Ending row number
        """
        # Apply formatting to risk columns
        if "Risk" in col_name_to_idx:
            risk_col = col_name_to_idx["Risk"]
            for row in range(start_row, end_row + 1):
                cell = ws.cell(row=row, column=risk_col)
                if cell.value in self.risk_colors:
                    cell.fill = PatternFill(
                        start_color=self.risk_colors[cell.value],
                        end_color=self.risk_colors[cell.value],
                        fill_type="solid"
                    )
                    cell.font = Font(bold=True)
        
        # Apply formatting to CVSS categories
        for cat_col in ["CVSS Category", "EPSS Category", "VPR Category"]:
            if cat_col in col_name_to_idx:
                col_idx = col_name_to_idx[cat_col]
                for row in range(start_row, end_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value in self.risk_colors:
                        cell.fill = PatternFill(
                            start_color=self.risk_colors[cell.value],
                            end_color=self.risk_colors[cell.value],
                            fill_type="solid"
                        )
        
        # Apply formatting to KEV Listed column
        if "KEV Listed" in col_name_to_idx:
            kev_col = col_name_to_idx["KEV Listed"]
            for row in range(start_row, end_row + 1):
                cell = ws.cell(row=row, column=kev_col)
                if cell.value == "Yes":
                    cell.fill = PatternFill(
                        start_color=self.risk_colors["Critical"],
                        end_color=self.risk_colors["Critical"],
                        fill_type="solid"
                    )
                    cell.font = Font(bold=True)
    
    def coordinate_from_string(self, cell_string):
        """
        Extract row and column from cell reference.
        
        Args:
            cell_string: Cell reference (e.g., "A1")
            
        Returns:
            Tuple of (column_string, row_number)
        """
        # Utility function to split cell references
        from openpyxl.utils.cell import coordinate_from_string
        return coordinate_from_string(cell_string)