"""Bucket Details sheet generator."""
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from va_tool.reporting.sheets.base import BaseSheetGenerator

class BucketDetailsSheetGenerator(BaseSheetGenerator):
    """Generator for the 3.4 Bucket Details sheet."""
    def __init__(self):
        super().__init__(title="3.4 Bucket Details")
        self.header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        self.header_font = Font(bold=True, color="003366")
        self.bucket_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    def generate(self, wb, df=None, **kwargs):
        ws = super().generate(wb)
        ws.title = "3.4 Bucket Details"
        #ws.sheet_properties.tabColor = "B7DEE8"
        
        ws.cell(row=1, column=1, value="Bucket-wise Vulnerability Details").font = Font(size=14, bold=True)
        ws.merge_cells('A1:H1')
        ws.cell(row=2, column=1, value="This sheet lists all vulnerabilities grouped by their assigned bucket, allowing detailed review and filtering.").font = Font(italic=True)
        ws.merge_cells('A2:H2')
        
        if df is None or 'Bucket' not in df.columns:
            ws.cell(row=4, column=1, value="No bucket data available.")
            return ws
        
        # Prepare columns to show
        columns = ["Bucket", "CVE", "Name", "Risk", "Host", "KEV Listed", "CVSS v3.0 Base Score", "EPSS Score"]
        col_map = {col: col for col in columns if col in df.columns}
        # Fallback: show all columns if some are missing
        if len(col_map) < len(columns):
            columns = [c for c in columns if c in df.columns]
            col_map = {col: col for col in columns}
        
        # Combine entries with same Name and Host, merge CVEs
        if 'Name' in df.columns and 'Host' in df.columns and 'CVE' in df.columns:
            agg_dict = {col: 'first' for col in df.columns if col not in ['CVE']}
            agg_dict['CVE'] = lambda x: ','.join(sorted(set(str(c) for c in x if pd.notna(c) and str(c).strip() != '')))
            df = df.groupby(['Name', 'Host'], as_index=False).agg(agg_dict)
        
        # Sort by Bucket, then Risk, then Name
        df_sorted = df.sort_values(["Bucket", "Risk", "Name"]).copy()
        
        # Write headers
        for idx, col in enumerate(columns, 1):
            cell = ws.cell(row=4, column=idx, value=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border
            ws.column_dimensions[get_column_letter(idx)].width = 22 if col=="Name" else 16
        
        # Write data, grouping by bucket
        row = 5
        for bucket, group in df_sorted.groupby("Bucket"):
            # Bucket header row
            ws.cell(row=row, column=1, value=bucket).font = Font(bold=True, color="006100")
            ws.cell(row=row, column=1).fill = self.bucket_fill
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
            row += 1
            # Write vulnerabilities for this bucket
            for _, vuln in group.iterrows():
                for idx, col in enumerate(columns, 1):
                    val = vuln.get(col, "")
                    ws.cell(row=row, column=idx, value=val)
                    ws.cell(row=row, column=idx).border = self.border
                row += 1
            # Add a blank row after each bucket
            row += 1
        return ws
