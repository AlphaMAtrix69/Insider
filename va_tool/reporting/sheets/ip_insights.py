from va_tool.reporting.sheets.base import BaseSheetGenerator
from va_tool.processing.scoring import calculate_exploitability_score
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart3D, Reference, BarChart, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import pandas as pd
import os
import numpy as np

class IPInsightsSheetGenerator(BaseSheetGenerator):
    """Generator for the IP Insights sheet."""
    
    def __init__(self):
        """Initialize the generator."""
        super().__init__(title="3.1 IP Insights")
    
    def generate(self, wb, df=None, output_dir=None, **kwargs):
        """
        Generate the IP Insights sheet.
        
        Args:
            wb: Excel workbook
            df: DataFrame with vulnerability data
            output_dir: Directory to save chart images
            **kwargs: Additional arguments
        
        Returns:
            The worksheet
        """
        self.logger.info("Generating IP Insights sheet")
        ws = super().generate(wb)
        
        # Add title
        self.add_title(ws, "IP Insights", font_size=14, merge_range='A1:L1')
        
        # Add introduction
        intro = ws.cell(row=3, column=1, value="This sheet provides a detailed breakdown of vulnerabilities per IP, facilitating operational planning and remediation prioritization.")
        intro.font = Font(italic=True)
        ws.merge_cells('A3:N3')
        
        # Section divider
        #ws.cell(row=4, column=1, value="="*60)
        
        # Add summary statistics at the top
        stats_row = 5
        if df is not None and 'Host' in df.columns:
            ip_data = self.process_ip_data(df)
            
            # Summary statistics
            total_hosts = len(set([ip["IP Address"] for ip in ip_data]))
            total_vulns = sum(ip["Total Vulnerabilities"] for ip in ip_data)
            ws.cell(row=stats_row, column=1, value=f"Total Hosts: {total_hosts}")
            ws.cell(row=stats_row, column=3, value=f"Total Vulnerabilities: {total_vulns}")
            
            # Top 10 hosts by vulnerabilities
            #top10 = sorted(ip_data, key=lambda x: x["Total Vulnerabilities"], reverse=True)[:10]
            #ws.cell(row=stats_row+2, column=1, value="Top 10 Hosts by Vulnerabilities:")
            #for i, ip_info in enumerate(top10, stats_row+3):
            #    ws.cell(row=i, column=1, value=ip_info["IP Address"])
            #    ws.cell(row=i, column=2, value=ip_info["Total Vulnerabilities"])
            
            # Section divider
            #ws.cell(row=stats_row+14, column=1, value="="*60)
            
            # Main table headers and data
            headers = [
                "IP Address", "Hostname", "Total Vulnerabilities", 
                "Critical", "High", "Medium", "Low", 
                "KEV Count", "Exploitability Score", "Most Common Category",
                "Last Scan Date"
            ]
            self.write_headers(ws, headers, row=stats_row+8)
            
            for row_idx, ip_info in enumerate(ip_data, stats_row+9):
                for col_idx, header in enumerate(headers, 1):
                    value = ip_info.get(header, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Apply conditional formatting
                    if header in ["Critical", "High", "Medium", "Low"] and value > 0:
                        cell.fill = PatternFill(
                            start_color=self.risk_colors[header],
                            end_color=self.risk_colors[header],
                            fill_type="solid"
                        )
                    
                    if header == "KEV Count" and value > 0:
                        cell.fill = PatternFill(
                            start_color=self.risk_colors["Critical"],
                            end_color=self.risk_colors["Critical"],
                            fill_type="solid"
                        )
                    
                    if header == "Exploitability Score":
                        if value >= 20:
                            cell.fill = PatternFill(start_color=self.risk_colors["Critical"], end_color=self.risk_colors["Critical"], fill_type="solid")
                        elif value >= 10:
                            cell.fill = PatternFill(start_color=self.risk_colors["High"], end_color=self.risk_colors["High"], fill_type="solid")
                        elif value >= 5:
                            cell.fill = PatternFill(start_color=self.risk_colors["Medium"], end_color=self.risk_colors["Medium"], fill_type="solid")
                    
                    # Add border to all cells
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            # Section divider before visualizations
            ws.cell(row=row_idx+2, column=1, value="="*60)
            
            # Visualizations
            if output_dir:
                try:
                    self.generate_3d_bar_chart(ws, ip_data, output_dir)
                except Exception as e:
                    self.logger.error(f"Error generating Excel 3D chart: {str(e)}")
                    self.generate_static_bar_chart(ws, ip_data, output_dir)
            
            # Add summary statistics (if any additional)
            self.add_summary_statistics(ws, ip_data)
        else:
            ws['A5'] = "No IP data available. Ensure the dataset contains a 'Host' column."
        
        return ws
    
    def process_ip_data(self, df):
        """Process IP data to create a comprehensive overview."""
        ip_data = []
        
        # Get unique hosts
        unique_hosts = df['Host'].unique()
        
        for host in unique_hosts:
            host_df = df[df['Host'] == host]
            
            # Count vulnerabilities by severity
            severity_counts = {
                "Critical": len(host_df[host_df['Risk'] == 'Critical']),
                "High": len(host_df[host_df['Risk'] == 'High']),
                "Medium": len(host_df[host_df['Risk'] == 'Medium']),
                "Low": len(host_df[host_df['Risk'] == 'Low'])
            }
            
            # Count KEV listed vulnerabilities
            kev_count = len(host_df[host_df.get('KEV Listed', '') == 'Yes'])
            
            # Calculate exploitability score (if data is available)
            exploitability_score = 0
            if 'Exploitability Score' in host_df.columns:
                exploitability_score = host_df['Exploitability Score'].sum()
            else:
                # Create a rough estimate from other factors
                exploitability_score = (
                    severity_counts["Critical"] * 4 +
                    severity_counts["High"] * 3 +
                    severity_counts["Medium"] * 2 +
                    severity_counts["Low"] * 1 +
                    kev_count * 3
                )
            
            # Determine most common vulnerability category
            most_common_category = "Unknown"
            if 'Bucket' in host_df.columns:
                # Split multi-category entries and count each occurrence
                all_categories = []
                for categories in host_df['Bucket'].dropna():
                    all_categories.extend([c.strip() for c in str(categories).split(',')])
                
                if all_categories:
                    category_counts = pd.Series(all_categories).value_counts()
                    most_common_category = category_counts.index[0] if not category_counts.empty else "Unknown"
            
            # Get last scan date (using current date as placeholder)
            last_scan_date = "N/A"  # In a real environment, you would use actual scan dates
            
            ip_data.append({
                "IP Address": host,
                "Total Vulnerabilities": len(host_df),
                "Critical": severity_counts["Critical"],
                "High": severity_counts["High"],
                "Medium": severity_counts["Medium"],
                "Low": severity_counts["Low"],
                "KEV Count": kev_count,
                "Exploitability Score": round(exploitability_score, 2),
                "Most Common Category": most_common_category,
                "Last Scan Date": last_scan_date
            })
        
        # Sort by total vulnerabilities in descending order
        ip_data.sort(key=lambda x: x["Total Vulnerabilities"], reverse=True)
        
        return ip_data
    
    def generate_3d_bar_chart(self, ws, ip_data, output_dir):
        """Generate a 3D bar chart of vulnerability distributions."""
        # Limit to top 10 IPs for better visualization
        top_ips = ip_data[:10]
        
        # Create chart data in worksheet
        data_start_row = 20
        
        # Add chart title
        chart_title = ws.cell(row=4, column=13, value="Top 10 Hosts by Vulnerability Count")
        chart_title.font = Font(size=12, bold=True, color=self.kpmg_dark_blue)
        
        # Write headers - these become the series titles
        headers = ["IP Address", "Critical", "High", "Medium", "Low"]
        for col, header in enumerate(headers, 13):
            ws.cell(row=5, column=col, value=header)
        
        # Write data
        for row, ip in enumerate(top_ips, 6):
            ws.cell(row=row, column=13, value=ip["IP Address"])
            ws.cell(row=row, column=14, value=ip["Critical"])
            ws.cell(row=row, column=15, value=ip["High"])
            ws.cell(row=row, column=16, value=ip["Medium"])
            ws.cell(row=row, column=17, value=ip["Low"])
        
        # Create 3D bar chart
        chart = BarChart3D()
        chart.title = "Vulnerability Distribution by Host"
        chart.style = 10
        chart.x_axis.title = "Host"
        chart.y_axis.title = "Vulnerability Count"
        chart.legend.position = 'r'
        
        # Create data references
        data = Reference(
            ws, 
            min_col=14,  # Start from column B (Critical)
            max_col=17,  # End at column E (Low)
            min_row=5,  # Include header row for series names
            max_row=5 + len(top_ips)
        )
        
        # Add data with series titles from the header row
        chart.add_data(data, titles_from_data=True)
        
        # Set categories
        cats = Reference(
            ws, 
            min_col=13, 
            max_col=13, 
            min_row=6,  # Skip header row for categories
            max_row=5 + len(top_ips)
        )
        chart.set_categories(cats)
        
        # Add data labels
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        
        # Set chart size and position
        chart.height = 13
        chart.width = 18
        
        # Add chart to worksheet
        ws.add_chart(chart, "L1")
    
    def generate_static_bar_chart(self, ws, ip_data, output_dir):
        """Generate a static bar chart image as a fallback."""
        # PNG output removed: do nothing
        self.logger.info("Static bar chart PNG output disabled.")
        return ws
    
    def add_summary_statistics(self, ws, ip_data):
        """Add summary statistics for operational planning."""
        # Calculate statistics
        total_hosts = len(ip_data)
        vulnerable_hosts = sum(1 for ip in ip_data if sum([ip["Critical"], ip["High"], ip["Medium"], ip["Low"]]) > 0)
        critical_hosts = sum(1 for ip in ip_data if ip["Critical"] > 0)
        high_hosts = sum(1 for ip in ip_data if ip["High"] > 0)
        
        # Add statistics section
        stats_row = 6
        
        # Add section title
        stats_title = ws.cell(row=stats_row, column=1, value="Operational Planning Statistics")
        stats_title.font = Font(size=12, bold=True, color=self.kpmg_blue)
        ws.merge_cells(start_row=stats_row, start_column=1, end_row=stats_row, end_column=6)
        
        # Add statistics data
        stats = [
            ["Total Hosts Scanned", total_hosts],
            ["Hosts with Vulnerabilities", vulnerable_hosts, f"{vulnerable_hosts/total_hosts*100:.1f}%" if total_hosts > 0 else "0%"],
            ["Hosts with Critical Vulnerabilities", critical_hosts, f"{critical_hosts/total_hosts*100:.1f}%" if total_hosts > 0 else "0%"],
            ["Hosts with High Vulnerabilities", high_hosts, f"{high_hosts/total_hosts*100:.1f}%" if total_hosts > 0 else "0%"],
            ["Top 5 Hosts Need Remediation", ", ".join([ip["IP Address"] for ip in ip_data[:5]])],
        ]
        
        ws.column_dimensions['A'].width = 20

        # Write statistics
        for i, stat in enumerate(stats, 1):
            row = stats_row + i
            
            # Label
            label_cell = ws.cell(row=row, column=1, value=stat[0])
            label_cell.font = Font(bold=True)
            label_cell.fill = PatternFill(start_color=self.kpmg_light_blue, end_color=self.kpmg_light_blue, fill_type="solid")
            label_cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Value
            value_cell = ws.cell(row=row, column=2, value=stat[1])
            value_cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Percentage (if available)
            if len(stat) > 2:
                percentage_cell = ws.cell(row=row, column=3, value=stat[2])
                percentage_cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # For the top hosts list, merge cells for better display
            if i == 5:  # Top 5 hosts
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
                value_cell.alignment = Alignment(horizontal="left")