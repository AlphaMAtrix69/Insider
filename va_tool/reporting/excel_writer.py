"""Common Excel writer functionality for report generation."""

import os
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from va_tool.utils import get_logger, style_header_cell, write_df_to_sheet, format_datetime, ensure_dir_exists

logger = get_logger()


def create_excel_workbook():
    """Create a new Excel workbook."""
    wb = Workbook()
    return wb


def save_excel_workbook(wb, output_dir, base_name="Vulnerability_Analysis"):
    """
    Save an Excel workbook with a timestamp.
    
    Args:
        wb: Workbook to save
        output_dir: Directory to save the workbook
        base_name: Base file name
    
    Returns:
        Path to the saved workbook
    """
    # Create timestamp
    timestamp = format_datetime(datetime.datetime.now())
    
    # Create directory if it doesn't exist
    ensure_dir_exists(output_dir)
    
    # Create file path
    output_path = os.path.join(output_dir, f"{base_name}_{timestamp}.xlsx")
    
    # Save workbook
    try:
        wb.save(output_path)
        logger.info(f"Excel workbook saved to {output_path}")
        return output_path
    except Exception as e:
        logger.error(f"Error saving Excel workbook: {str(e)}")
        return None


def add_title(ws, title, font_size=16, bold=True, merge_range="A1:H1"):
    """Add a title to a worksheet."""
    from openpyxl.styles import Font
    
    ws['A1'] = title
    ws['A1'].font = Font(size=font_size, bold=bold)
    ws.merge_cells(merge_range)
    return ws


def add_section_title(ws, title, cell="A3", bold=True):
    """Add a section title to a worksheet."""
    from openpyxl.styles import Font
    
    ws[cell] = title
    ws[cell].font = Font(bold=bold)
    return ws


def create_basic_report(df, output_dir, filename="Vulnerability_Analysis"):
    """
    Create a basic Excel report with the DataFrame as a single sheet.
    
    Args:
        df: DataFrame to write to Excel
        output_dir: Directory to save the Excel file
        filename: Base filename
    
    Returns:
        Path to the saved Excel file
    """
    logger.info("Creating basic Excel report")
    
    # Create timestamp
    timestamp = format_datetime(datetime.datetime.now())
    
    # Create directory if it doesn't exist
    ensure_dir_exists(output_dir)
    
    # Create file path
    output_path = os.path.join(output_dir, f"{filename}_{timestamp}.xlsx")
    
    # Save DataFrame to Excel
    try:
        df.to_excel(output_path, sheet_name="Vulnerability Data", index=False)
        logger.info(f"Basic Excel report saved to {output_path}")
        return output_path
    except Exception as e:
        logger.error(f"Error saving basic Excel report: {str(e)}")
        return None


def write_full_excel_report(original_df, processed_df, check_needed_df, analyzed_df, 
                           results_data, output_dir, filename="Vulnerability_Analysis"):
    """
    Write a full Excel report with multiple sheets.
    
    Args:
        original_df: Original vulnerability DataFrame
        processed_df: Processed vulnerability DataFrame
        check_needed_df: Vulnerabilities that need checking
        analyzed_df: Analyzed vulnerability DataFrame
        results_data: Dictionary with analysis results
        output_dir: Directory to save the Excel file
        filename: Base filename
    
    Returns:
        Path to the saved Excel file
    """
    logger.info("Creating full Excel report")
    
    # Create timestamp
    timestamp = format_datetime(datetime.datetime.now())
    
    # Create directory if it doesn't exist
    ensure_dir_exists(output_dir)
    
    # Create file path
    output_path = os.path.join(output_dir, f"{filename}_{timestamp}.xlsx")
    
    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Write main data sheets
            processed_df.to_excel(writer, sheet_name="Updated Data", index=False)
            check_needed_df.to_excel(writer, sheet_name="Checks Needed", index=False)
            
            # Write summary sheets
            risk_counts = pd.DataFrame(results_data.get("risk_counts", []))
            if not risk_counts.empty:
                risk_counts.to_excel(writer, sheet_name="Risk Summary", index=False)
                
            top_cves = pd.DataFrame(results_data.get("top_cves", []))
            if not top_cves.empty:
                top_cves.to_excel(writer, sheet_name="Top CVEs", index=False)
                
            insights = pd.DataFrame(results_data.get("vulnerability_insights", []))
            if not insights.empty:
                insights.to_excel(writer, sheet_name="Vulnerability Insights", index=False)
                
            ip_insights = pd.DataFrame(results_data.get("ip_vuln_counts", []))
            if not ip_insights.empty:
                ip_insights.to_excel(writer, sheet_name="Per IP Insights", index=False)
                
            bucket_summary = pd.DataFrame(results_data.get("bucket_summary", []))
            if not bucket_summary.empty:
                bucket_summary.to_excel(writer, sheet_name="Bucket Summary", index=False)
                
            patch_summary = pd.DataFrame(results_data.get("patch_summary", []))
            if not patch_summary.empty:
                patch_summary.to_excel(writer, sheet_name="Patch Summary", index=False)
                
            scoring_summary = pd.DataFrame(results_data.get("scoring_summary", []))
            if not scoring_summary.empty:
                scoring_summary.to_excel(writer, sheet_name="Scoring Summary", index=False)
                
            common_critical = pd.DataFrame(results_data.get("common_critical", []))
            if not common_critical.empty:
                common_critical.to_excel(writer, sheet_name="Common Priority", index=False)
                
            # Add SEOL data if available
            seol_items = analyzed_df[analyzed_df["Name"].str.contains("SEoL", case=False, na=False)]
            if not seol_items.empty:
                seol_items.to_excel(writer, sheet_name="SEOL", index=False)
                
                seol_summary = pd.DataFrame(results_data.get("seol_summary", []))
                if not seol_summary.empty:
                    seol_summary.to_excel(writer, sheet_name="SEOL IP Summary", index=False)
        
        logger.info(f"Full Excel report saved to {output_path}")
        return output_path
    except Exception as e:
        logger.error(f"Error saving full Excel report: {str(e)}")
        return None