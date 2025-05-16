"""Top 10 CVEs sheet generator."""

import pandas as pd
import matplotlib.pyplot as plt
import os
from collections import Counter
from openpyxl.drawing.image import Image

from va_tool.reporting.sheets.base import BaseSheetGenerator
from va_tool.utils import ensure_dir_exists


class Top10CVEsSheetGenerator(BaseSheetGenerator):
    """Generator for the Top 10 CVEs sheet."""
    
    def __init__(self):
        """Initialize the generator."""
        super().__init__(title="5.2 Top 10 CVEs")
    
    def generate(self, wb, df=None, output_dir=None, **kwargs):
        """
        Generate the Top 10 CVEs sheet with bar chart and additional CVE details.
        
        Args:
            wb: Excel workbook
            df: DataFrame with vulnerability data
            output_dir: Directory to save chart images
            **kwargs: Additional arguments
        
        Returns:
            The worksheet
        """
        self.logger.info("Generating Top 10 CVEs sheet")
        ws = super().generate(wb)
        
        # Add title
        self.add_title(
            ws, "Top 10 CVEs", 
            font_size=14
        )
        
        try:
            if df is not None and 'CVE' in df.columns:
                # Extract CVEs and relevant columns
                df_cves = df[["CVE", "Name", "Description", "Host", "Solution"]].dropna(subset=["CVE"])
                
                # Process CVEs and collect details
                all_cves, cve_details = self.extract_cve_details(df_cves)
                
                # Count CVEs and get top 10
                cve_counts = Counter(all_cves)
                top_10 = cve_counts.most_common(10)
                
                if top_10:
                    # Define headers
                    headers = ["CVE", "Count", "Name", "Description", "Hosts", "Solutions"]
                    self.write_headers(ws, headers, row=3)
                    
                    # Write top 10 data
                    for i, (cve, count) in enumerate(top_10, start=4):
                        ws[f"A{i}"] = cve
                        ws[f"B{i}"] = count
                        ws[f"C{i}"] = cve_details.get(cve, {}).get("name", "N/A")
                        ws[f"D{i}"] = cve_details.get(cve, {}).get("description", "N/A")
                        
                        hosts = ", ".join(sorted(cve_details.get(cve, {}).get("hosts", [])))
                        ws[f"E{i}"] = hosts if hosts else "N/A"
                        
                        solutions = ", ".join(sorted(cve_details.get(cve, {}).get("solutions", [])))
                        ws[f"F{i}"] = solutions if solutions else "N/A"
                    
                    # Adjust column widths
                    ws.column_dimensions["A"].width = 20  # CVE
                    ws.column_dimensions["B"].width = 10  # Count
                    ws.column_dimensions["C"].width = 50  # Name
                    ws.column_dimensions["D"].width = 80  # Description
                    ws.column_dimensions["E"].width = 50  # Hosts
                    ws.column_dimensions["F"].width = 80  # Solutions
                    
                    # Generate chart if output_dir is provided
                    if output_dir:
                        self.generate_bar_chart(top_10, output_dir, ws)
                else:
                    ws['A3'] = "No CVEs found in the data."
            else:
                ws['A3'] = "No vulnerability data or 'CVE' column available."
        
        except Exception as e:
            self.logger.error(f"Error generating Top 10 CVEs sheet: {str(e)}")
            ws['A3'] = f"Error generating Top 10 CVEs: {str(e)}"
        
        return ws
    
    def extract_cve_details(self, df_cves):
        """Extract CVE details from the dataframe."""
        all_cves = []
        cve_details = {}
        
        # Process each row
        for _, row in df_cves.iterrows():
            # Split multiple CVEs in one field
            cves = [cve.strip() for cve in str(row["CVE"]).split(',') if cve.strip().startswith('CVE')]
            
            for cve in cves:
                all_cves.append(cve)
                
                # Initialize details dictionary if this is the first time seeing this CVE
                if cve not in cve_details:
                    cve_details[cve] = {
                        "name": row["Name"] if pd.notna(row["Name"]) else "N/A",
                        "description": row["Description"] if pd.notna(row["Description"]) else "N/A",
                        "hosts": set(),
                        "solutions": set()
                    }
                
                # Add host and solution if available
                if pd.notna(row["Host"]):
                    cve_details[cve]["hosts"].add(str(row["Host"]))
                if pd.notna(row["Solution"]):
                    cve_details[cve]["solutions"].add(str(row["Solution"]))
        
        return all_cves, cve_details
    
    def generate_bar_chart(self, top_10, output_dir, ws):
        """Generate a bar chart for the top 10 CVEs."""
        if not top_10:
            return
        
        # Extract data for chart
        cve_names, counts = zip(*top_10)
        
        plt.figure(figsize=(12, 6))
        bars = plt.bar(cve_names, counts, color='blue')
        
        # Calculate maximum count for layout
        max_count = max(counts) if counts else 1
        plt.ylim(0, max_count * 1.2)
        
        # Add labels and title
        plt.xlabel('CVE', labelpad=15, fontsize=12, fontweight='bold')
        plt.ylabel('Occurrences', labelpad=15, fontsize=12, fontweight='bold')
        plt.title('TOP 10 CVEs', pad=20, fontsize=14, fontweight='bold')
        plt.xticks(rotation=45, ha='right')
        
        # Add value labels on bars
        for bar in bars:
            yval = bar.get_height()
            plt.text(bar.get_x() + bar.get_width() / 2, yval + (max_count * 0.02), int(yval),
                    ha='center', va='bottom', fontsize=9)
        
        # Layout adjustments
        plt.subplots_adjust(left=0.1, right=0.95, top=0.9, bottom=0.25)
        plt.tight_layout()
        
        # Save chart
        ensure_dir_exists(output_dir)
        chart_path = os.path.join(output_dir, "cve_chart.png")
        plt.savefig(chart_path)
        plt.close()
        
        # Add chart to Excel
        try:
            img = Image(chart_path)
            img.width = 720
            img.height = 400
            ws.add_image(img, "A15")  # Place chart below the table
            self.logger.info(f"Added CVE chart to sheet: {chart_path}")
        except Exception as e:
            self.logger.error(f"Error adding CVE chart: {str(e)}")
            ws['A15'] = "Error adding chart image"