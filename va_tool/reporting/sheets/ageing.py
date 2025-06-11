"""Ageing of Vulnerabilities sheet generator."""

from va_tool.reporting.sheets.base import BaseSheetGenerator
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart3D, Reference
from openpyxl.chart.label import DataLabelList


class AgeingSheetGenerator(BaseSheetGenerator):
    """Generator for the Ageing of Vulnerabilities sheet."""
    
    def __init__(self):
        """Initialize the generator."""
        super().__init__(title="1.2 Ageing of Vulnerability")
        
        # KPMG color scheme
        self.kpmg_blue = "00338D"
        self.kpmg_light_blue = "DCE6F1"
        self.kpmg_dark_blue = "005EB8"
        self.kpmg_medium_blue = "0091DA"
        self.kpmg_accent_blue = "00A3E0"
        
        # Risk colors
        self.risk_colors = {
            "Critical": "FF7D7D",  # Darker pastel red
            "High": "FFB366",      # Darker pastel orange
            "Medium": "FFDA66",    # Darker pastel yellow
            "Low": "99CC99"        # Darker pastel green
        }
        
        # Age ranges colors - as vulnerabilities age, they become more critical
        self.age_colors = {
            "0-30 days": "92D050",     # Light green
            "31-90 days": "FFFF00",    # Yellow
            "91-180 days": "FFC000",   # Orange
            "181-365 days": "FF0000",  # Red
            "Over 1 year": "7030A0"    # Purple
        }
    
    def generate(self, wb, df=None, **kwargs):
        """
        Generate the Ageing of Vulnerabilities sheet.
        
        Args:
            wb: Excel workbook
            df: DataFrame with vulnerability data
            **kwargs: Additional arguments
        
        Returns:
            The worksheet
        """
        self.logger.info("Generating Ageing of Vulnerabilities sheet")
        ws = super().generate(wb)
        
        # Apply KPMG styling
        self.apply_kpmg_styling(ws)
        
        # Add title with enhanced styling
        self.add_title(
            ws, "Ageing of Vulnerabilities", 
            font_size=14, merge_range='A1:H1'
        )
        
        # Add age distribution summary
        self.add_age_distribution_summary(ws, df)
        
        # Tenable API integration is not available, so this part is commented out
        # note_cell = ws['A4']
        # note_cell.value = "Note: Plugin Initial Release Date and Plugin Updated Date will be populated from Tenable API once integrated."
        # note_cell.font = Font(italic=True, color=self.kpmg_dark_blue)
        # ws.merge_cells('A4:H4')
        
        # Check if we have the data
        if df is not None:
            # Define the columns to display
            columns = [
                "Plugin ID", "CVE", "Host", "Name", "Risk", 
                "CVE Published Date", "Days After Discovery"
            ]
            
            # Tenable API columns are not available, so this part is commented out
            # extended_columns = columns + ["Plugin Initial Release Date", "Plugin Updated Date"]
            
            # Add two empty rows for better spacing between summary and main data
            # for row in [6, 7]:
            #     for col in range(1, len(extended_columns) + 1):
            #         ws.cell(row=row, column=col, value="")
            # 
            # # Add headers with enhanced styling (moved to row 8)
            self.add_headers(ws, columns, row=11)
            
            # Check if required columns exist
            if all(col in df.columns for col in columns):
                # Sort the DataFrame by "Days After Discovery" (descending)
                subset_df = df[columns].copy()
                subset_df = subset_df.sort_values(by="Days After Discovery", ascending=False)
                
                # Write data with enhanced styling (starting from row 9)
                for row_idx, (_, row) in enumerate(subset_df.iterrows(), 12):
                    for col_idx, col in enumerate(columns, 1):
                        value = row.get(col, '')
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        
                        # Apply conditional formatting based on value
                        self.apply_conditional_formatting(cell, col, value)
                    # Tenable API integration is not available, so this part is commented out
                    # Add empty cells for Tenable API data (to be filled later)
                    # ws.cell(row=row_idx, column=len(columns) + 1, value="N/A")
                    # ws.cell(row=row_idx, column=len(columns) + 2, value="N/A")
            else:
                ws['A9'] = "Required columns not found in data."
        else:
            ws['A8'] = "No vulnerability data available."
        
        return ws
    
    def apply_kpmg_styling(self, ws):
        """Apply KPMG styling to the worksheet."""
        # Set title row styling
        title_cell = ws['A1']
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color=self.kpmg_blue, end_color=self.kpmg_blue, fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set row heights
        ws.row_dimensions[1].height = 30
        
        # Set column widths
        ws.column_dimensions['A'].width = 15  # Plugin ID
        ws.column_dimensions['B'].width = 18  # CVE
        ws.column_dimensions['C'].width = 15  # Host
        ws.column_dimensions['D'].width = 40  # Name
        ws.column_dimensions['E'].width = 12  # Risk
        ws.column_dimensions['F'].width = 20  # CVE Published Date
        ws.column_dimensions['G'].width = 20  # Days After Discovery
        #ws.column_dimensions['I'].width = 25  # Plugin Updated Date
    
    def add_headers(self, ws, headers, row=5):
        """Add headers with enhanced KPMG styling."""
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=i, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.kpmg_dark_blue, end_color=self.kpmg_dark_blue, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Add border
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    def apply_conditional_formatting(self, cell, column, value):
        """Apply conditional formatting based on column and value."""
        # Add standard border
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format Risk column
        if column == "Risk" and value in self.risk_colors:
            cell.fill = PatternFill(
                start_color=self.risk_colors[value],
                end_color=self.risk_colors[value],
                fill_type="solid"
            )
            cell.font = Font(bold=True)
        
        # Format Days After Discovery column
        elif column == "Days After Discovery" and value is not None:
            try:
                days = int(value)
                # Apply color based on age ranges
                if days <= 30:
                    color = self.age_colors["0-30 days"]
                elif days <= 90:
                    color = self.age_colors["31-90 days"]
                elif days <= 180:
                    color = self.age_colors["91-180 days"]
                elif days <= 365:
                    color = self.age_colors["181-365 days"]
                else:
                    color = self.age_colors["Over 1 year"]
                
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                # Make text bold for older vulnerabilities
                if days > 180:
                    cell.font = Font(bold=True)
            except (ValueError, TypeError):
                pass
    
    def add_age_distribution_summary(self, ws, df):
        """Add a summary section for vulnerability age distribution."""
        # Add section title
        summary_title = ws.cell(row=3, column=1, value="Vulnerability Age Distribution")
        summary_title.font = Font(size=12, bold=True, color=self.kpmg_blue)
        ws.merge_cells('A3:D3')

        # Define age buckets
        age_buckets = {
            "0-30 days": (0, 30),
            "31-90 days": (31, 90),
            "91-180 days": (91, 180),
            "181-365 days": (181, 365),
            "Over 1 year": (366, float('inf'))
        }

        # Calculate age distribution if data available
        age_counts = {}
        total_count = 0

        if df is not None and "Days After Discovery" in df.columns:
            for bucket, (min_age, max_age) in age_buckets.items():
                count = len(df[(df["Days After Discovery"] >= min_age) & 
                              (df["Days After Discovery"] <= max_age)])
                age_counts[bucket] = count
                total_count += count
        else:
            # Sample data if no data available
            age_counts = {
                "0-30 days": 1,
                "31-90 days": 1,
                "91-180 days": 1,
                "181-365 days": 1,
                "Over 1 year": 1
            }
            total_count = sum(age_counts.values())

        # Add headers for age distribution table
        ws.cell(row=4, column=1, value="Age Range").font = Font(bold=True)
        ws.cell(row=4, column=2, value="Count").font = Font(bold=True)
        ws.cell(row=4, column=3, value="Percentage").font = Font(bold=True)

        # Style header cells
        for col in range(1, 4):
            cell = ws.cell(row=4, column=col)
            cell.fill = PatternFill(start_color=self.kpmg_light_blue, end_color=self.kpmg_light_blue, fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Write age distribution data
        for i, (age_range, count) in enumerate(age_counts.items(), 4):
            percentage = (count / total_count * 100) if total_count > 0 else 0

            age_cell = ws.cell(row=i, column=1, value=age_range)
            count_cell = ws.cell(row=i, column=2, value=count)
            percentage_cell = ws.cell(row=i, column=3, value=f"{percentage:.1f}%")

            # Style based on age range
            color = self.age_colors[age_range]
            age_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            count_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            percentage_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # Add borders
            for col in range(1, 4):
                cell = ws.cell(row=i, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

        # Add bar chart for age distribution
        chart = BarChart3D()
        chart.type = "col"
        chart.style = 34
        chart.title = "Vulnerability Age Distribution"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Age Range"

        # Add data
        data = Reference(ws, min_col=2, max_col=2, min_row=4, max_row=4+len(age_counts))
        cats = Reference(ws, min_col=1, max_col=1, min_row=5, max_row=4+len(age_counts))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Add data labels
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True

        # Set chart size and position for better proportions
        chart.height = 10
        chart.width = 13
        ws.add_chart(chart, "H1")