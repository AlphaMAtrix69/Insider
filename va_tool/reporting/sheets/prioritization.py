"""Prioritization Insights sheet generator."""

from va_tool.reporting.sheets.base import BaseSheetGenerator
from va_tool.utils import style_header_cell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class PrioritizationInsightsGenerator(BaseSheetGenerator):
    """Generator for the Prioritization Insights sheet."""
    
    def __init__(self):
        """Initialize the generator."""
        super().__init__(title="1. Prioritization Insights")
        
        # KPMG color scheme
        self.kpmg_blue = "00338D"
        self.kpmg_light_blue = "DCE6F1"
        self.kpmg_dark_blue = "005EB8"
        self.kpmg_medium_blue = "0091DA"
        self.kpmg_accent_blue = "00A3E0"
        
        # Risk colors
        self.risk_colors = {
            "Critical": "FF7D7D",  # Darker pastel red
            "High": "FFB366",      # Darker pastel orange
            "Medium": "FFDA66",    # Darker pastel yellow
            "Low": "99CC99"        # Darker pastel green
        }
    
    def generate(self, wb, results_data, original_df=None, processed_df=None, excluded_df=None, **kwargs):
        """
        Generate the Prioritization Insights sheet.
        
        Args:
            wb: Excel workbook
            results_data: Dictionary with analysis results
            original_df: Original vulnerability DataFrame
            processed_df: Processed vulnerability DataFrame
            excluded_df: Excluded entries DataFrame
            **kwargs: Additional arguments
        
        Returns:
            The worksheet
        """
        self.logger.info("Generating Prioritization Insights sheet")
        ws = super().generate(wb)
        
        # Style the worksheet
        self.apply_kpmg_theme(ws)
        
        # Add Raw Data section
        self.add_section_title(ws, "Nessus Raw Data count", cell="A1")
        self.add_raw_data_section(ws, original_df)
        # Add Excluded Entries section
        if excluded_df is not None and not excluded_df.empty:
            self.add_section_title(ws, "Excluded Entries (Informational, None, Check Needed)", cell="A6")
            self.add_excluded_entries_section(ws, excluded_df, start_row=7)
        
        # Add Revision Count section
        self.add_section_title(ws, "Revision Count", cell="H1")
        self.add_revision_count_section(ws, processed_df, excluded_df=excluded_df)
        
        # Add EPSS Risk Summary section
        self.add_section_title(ws, "EPSS Risk Summary", cell="H6")
        self.add_epss_summary(ws, processed_df)
        
        # Add VPR Risk Summary section
        self.add_section_title(ws, "VPR Risk Summary", cell="H10")
        self.add_vpr_summary(ws, processed_df)
        
        # Add KEV Risk Summary section
        self.add_section_title(ws, "KEV Risk Summary", cell="H14")
        self.add_kev_summary(ws, processed_df)
        
        # Add Metric table
        self.add_metric_table(ws, processed_df)
        # Add mini summary table below metrics summary
        self.add_common_priority_summary(ws, processed_df, start_row=25)
        
        # Add Prioritization Insights section
        self.add_prioritization_insights(ws, processed_df, results_data)
        
        # Add legend section with improved styling
        self.add_legend_section(ws)
        
        # Add summary bar chart at A17
        self.add_summary_bar_chart(ws)
        
        return ws
    
    def apply_kpmg_theme(self, ws):
        """Apply KPMG theme to the worksheet."""
        # Add a header with KPMG styling
        header = ws.cell(row=1, column=1, value="Prioritization Insights")
        header.font = Font(size=14, bold=True, color="FFFFFF")
        header.fill = PatternFill(start_color=self.kpmg_blue, end_color=self.kpmg_blue, fill_type="solid")
        header.alignment = Alignment(horizontal="center", vertical="center")
        
        # Merge header cells
        ws.merge_cells('A1:F1')
        
        # Set column widths
        for col in range(1, 25):
            col_letter = chr(64 + col) if col <= 26 else chr(64 + col // 26) + chr(64 + col % 26)
            ws.column_dimensions[col_letter].width = 15
        
        # Set specific column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['P'].width = 20
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
    
    def add_raw_data_section(self, ws, original_df):
        """Add raw data count section with KPMG styling (Critical, High, Medium, Low, None only)."""
        # Add headers
        headers = ["", "Critical", "High", "Medium", "Low", "None"]
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=i, value=header)
            style_header_cell(cell, bg_color=self.kpmg_light_blue)
            if header in self.risk_colors:
                cell.fill = PatternFill(start_color=self.risk_colors[header], end_color=self.risk_colors[header], fill_type="solid")
        ws['A4'] = "Total (Without None)"
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in range(2, 5):
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border
        # Add raw count values if available
        if original_df is not None:
            raw_counts = original_df["Risk"].value_counts()
            for i, risk in enumerate(["Critical", "High", "Medium", "Low", None ], 2):
                ws.cell(row=3, column=i, value=raw_counts.get(risk, 0))
            total_without_none = sum(raw_counts.get(risk, 0) for risk in ["Critical", "High", "Medium", "Low"])
            ws.cell(row=4, column=2, value=total_without_none)

    def add_revision_count_section(self, ws, processed_df, excluded_df=None):
        """Add revision count section with KPMG styling, including Informational and Check Needed from categorized df."""
        # Add headers
        revision_headers = ["", "Critical", "High", "Medium", "Low", "Informational", "N/A"]
        for i, header in enumerate(revision_headers, 8):
            cell = ws.cell(row=2, column=i, value=header)
            style_header_cell(cell, bg_color=self.kpmg_light_blue)
            if header in self.risk_colors:
                cell.fill = PatternFill(start_color=self.risk_colors[header], end_color=self.risk_colors[header], fill_type="solid")
        ws['H3'] = "Risk Rating:"
        ws['H4'] = "Total Without NA+Info"
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in range(2, 5):
            for col in range(8, 15):
                ws.cell(row=row, column=col).border = border
        # Add revision count values if available
        info_count = 0
        na_count = 0
        if excluded_df is not None:
            info_count = len(excluded_df[excluded_df["Risk"] == "Informational"])
            na_count = len(excluded_df[(excluded_df["Risk"] == "None") | (excluded_df["Risk"] == "Check Needed")])
        if processed_df is not None:
            revision_counts = processed_df["Risk"].value_counts()
            # Always show all columns, even if zero
            for i, risk in enumerate(["Critical", "High", "Medium", "Low"], 9):
                ws.cell(row=3, column=i, value=revision_counts.get(risk, 0))
            ws.cell(row=3, column=13, value=info_count)  # Informational
            ws.cell(row=3, column=14, value=na_count)    # N/A
            # Calculate total without Informational, Check Needed
            total_without_na = sum(revision_counts.get(risk, 0) for risk in ["Critical", "High", "Medium", "Low"])
            ws.cell(row=4, column=9, value=total_without_na)
    
    def add_epss_summary(self, ws, processed_df):
        """Add EPSS risk summary section with KPMG styling."""
        # Add header styling
        cell = ws.cell(row=6, column=8, value="EPSS Risk Summary")
        cell.font = Font(bold=True, color=self.kpmg_dark_blue)
        
        ws['H7'] = "EPSS Score Available"
        
        # Add border for EPSS cells
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(7, 9):
            for col in range(8, 13):
                ws.cell(row=row, column=col).border = border
        
        if processed_df is not None and 'EPSS Category' in processed_df.columns:
            epss_counts = processed_df['EPSS Category'].value_counts()
            total_epss = sum(epss_counts.get(risk, 0) for risk in ["Critical", "High", "Medium", "Low"])
            ws['H8'] = f"Total EPSS Score Available ({total_epss})"
            
            # Add EPSS counts by category
            for i, risk in enumerate(["Critical", "High", "Medium", "Low"], 9):
                cell = ws.cell(row=7, column=i, value=epss_counts.get(risk, 0))
                if risk in self.risk_colors:
                    cell.fill = PatternFill(start_color=self.risk_colors[risk], end_color=self.risk_colors[risk], fill_type="solid")
        else:
            ws['H8'] = "Total EPSS Score Available (0)"
    
    def add_vpr_summary(self, ws, processed_df):
        """Add VPR risk summary section with KPMG styling."""
        # Add header styling
        cell = ws.cell(row=10, column=8, value="VPR Risk Summary")
        cell.font = Font(bold=True, color=self.kpmg_dark_blue)
        
        ws['H11'] = "VPR Score Available"
        
        # Add border for VPR cells
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(11, 13):
            for col in range(8, 13):
                ws.cell(row=row, column=col).border = border
        
        if processed_df is not None and 'VPR Category' in processed_df.columns:
            vpr_counts = processed_df['VPR Category'].value_counts()
            total_vpr = sum(vpr_counts.get(risk, 0) for risk in ["Critical", "High", "Medium", "Low"])
            ws['H12'] = f"Total VPR Score Available ({total_vpr})"
            
            # Add VPR counts by category
            for i, risk in enumerate(["Critical", "High", "Medium", "Low"], 9):
                cell = ws.cell(row=11, column=i, value=vpr_counts.get(risk, 0))
                if risk in self.risk_colors:
                    cell.fill = PatternFill(start_color=self.risk_colors[risk], end_color=self.risk_colors[risk], fill_type="solid")
        else:
            ws['H12'] = "Total VPR Score Available (0)"
    
    def add_kev_summary(self, ws, processed_df):
        """Add KEV risk summary section with KPMG styling."""
        # Add header styling
        cell = ws.cell(row=14, column=8, value="KEV Risk Summary")
        cell.font = Font(bold=True, color=self.kpmg_dark_blue)
        
        ws['H15'] = "KEV Score Available"
        
        # Add border for KEV cells
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(15, 16):
            # Only apply to columns H (8), I (9), and J (10)
            # Leaving column G (7) empty
            for col in range(8, 11):
                ws.cell(row=row, column=col).border = border
        
        if processed_df is not None and 'KEV Listed' in processed_df.columns:
            kev_counts = processed_df['KEV Listed'].value_counts()
            yes_count = kev_counts.get('Yes', 0)
            ws.cell(row=15, column=9, value=yes_count)
            ws.cell(row=15, column=10, value=kev_counts.get('No', 0))
            
            # Highlight if KEVs are found with pastel colors for better readability
            if yes_count > 0:
                ws.cell(row=15, column=9).fill = PatternFill(
                    start_color="FF9999",  # Pastel red
                    end_color="FF9999", 
                    fill_type="solid"
                )
                
        # Make sure column G is left empty
        # Clear any content or formatting from column G
        ws.cell(row=15, column=7, value=None)
        ws.cell(row=15, column=7).fill = PatternFill(fill_type=None)

    def add_metric_table(self, ws, processed_df):
        """Add metric table with KPMG styling."""
        # Add header
        header_cell = ws.cell(row=19, column=8, value="Scoring Metrics Summary")
        header_cell.font = Font(size=12, bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color=self.kpmg_blue, end_color=self.kpmg_blue, fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('H19:L19')
        
        # Add column headers
        headers = ["Metric", "Critical Count", "High Count", "Medium Count", "Low Count"]
        for i, header in enumerate(headers, 8):
            cell = ws.cell(row=20, column=i, value=header)
            style_header_cell(cell, bg_color=self.kpmg_light_blue)
        
        # Add metric rows with styling
        metrics = ["CVSS", "EPSS", "VPR", "CISA KEV"]
        for i, metric in enumerate(metrics, 21):
            cell = ws.cell(row=i, column=8, value=metric)
            cell.fill = PatternFill(start_color=self.kpmg_light_accent, end_color=self.kpmg_light_accent, fill_type="solid")
            cell.font = Font(bold=True)
        
        # Add border for table
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(20, 25):
            for col in range(8, 13):
                ws.cell(row=row, column=col).border = border
        
        # Fill in metric counts if processed_df is available
        if processed_df is not None:
            for i, metric in enumerate(['CVSS Category', 'EPSS Category', 'VPR Category', 'KEV Listed'], 0):
                row_idx = 21 + i
                
                if metric == 'KEV Listed':
                    # For KEV, we need Yes/No counts
                    kev_counts = processed_df[metric].value_counts()
                    kev_yes = kev_counts.get('Yes', 0)
                    ws.cell(row=row_idx, column=9, value=kev_yes)
                    
                    # Highlight critical KEVs
                    if kev_yes > 0:
                        ws.cell(row=row_idx, column=9).fill = PatternFill(start_color=self.risk_colors["Critical"], end_color=self.risk_colors["Critical"], fill_type="solid")
                    
                    ws.cell(row=row_idx, column=10, value=0)
                    ws.cell(row=row_idx, column=11, value=0)
                    ws.cell(row=row_idx, column=12, value=0)
                else:
                    # For scoring metrics, count by category
                    counts = processed_df[metric].value_counts()
                    
                    for j, risk in enumerate(["Critical", "High", "Medium", "Low"], 9):
                        count = counts.get(risk, 0)
                        cell = ws.cell(row=row_idx, column=j, value=count)
                        
                        # Add conditional formatting for non-zero values
                        if count > 0 and risk in self.risk_colors:
                            cell.fill = PatternFill(start_color=self.risk_colors[risk], end_color=self.risk_colors[risk], fill_type="solid")
    
    def add_prioritization_insights(self, ws, processed_df, results_data):
        """Add prioritization insights section."""
        # Add insights section header
        header_cell = ws.cell(row=6, column=1, value="Key Prioritization Insights")
        header_cell.font = Font(size=12, bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color=self.kpmg_dark_blue, end_color=self.kpmg_dark_blue, fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A6:F6')
        
        # Determine insights
        insights = []
        
        # Get critical vulnerabilities count
        critical_count = 0
        if processed_df is not None and "Risk" in processed_df.columns:
            critical_count = len(processed_df[processed_df["Risk"] == "Critical"])
        
        # Get KEV listed vulnerability count
        kev_count = 0
        if processed_df is not None and "KEV Listed" in processed_df.columns:
            kev_count = len(processed_df[processed_df["KEV Listed"] == "Yes"])
        
        # Get common critical vulnerabilities
        common_critical_count = len(results_data.get("common_critical", []))
        
        # Get top vulnerability by bucket
        bucket_data = results_data.get("bucket_summary", [])
        top_bucket = None
        if bucket_data:
            top_bucket = sorted(bucket_data, key=lambda x: x.get("Count", 0), reverse=True)[0].get("Bucket")
        
        # Generate insights
        insights.append(f"• {critical_count} Critical vulnerabilities detected requiring immediate attention")
        insights.append(f"• {kev_count} vulnerabilities on CISA Known Exploited Vulnerabilities (KEV) catalog")
        insights.append(f"• {common_critical_count} vulnerabilities are rated critical by multiple scoring systems")
        
        if top_bucket:
            insights.append(f"• {top_bucket} vulnerabilities represent the largest category of issues")
        
        insights.append("• Prioritization Methodology:")
        insights.append("  1. Address all KEV-listed vulnerabilities first")
        insights.append("  2. Focus on vulnerabilities rated critical by multiple scoring systems")
        insights.append("  3. Prioritize vulnerabilities with high EPSS scores indicating likelihood of exploitation")
        insights.append("  4. Address vulnerabilities by risk severity (Critical → High → Medium → Low)")
        
        # Add border and styling for insights section
        border = Border(
            left=Side(style='medium', color=self.kpmg_dark_blue),
            right=Side(style='medium', color=self.kpmg_dark_blue),
            top=Side(style='medium', color=self.kpmg_dark_blue),
            bottom=Side(style='medium', color=self.kpmg_dark_blue)
        )
        
        # Background for insights box
        for row in range(6, 6 + len(insights) + 1):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                if row == 6:  # Header already styled
                    continue
                cell.fill = PatternFill(start_color=self.kpmg_light_blue, end_color=self.kpmg_light_blue, fill_type="solid")
                cell.border = border
        
        # Write insights
        for i, insight in enumerate(insights, 7):
            cell = ws.cell(row=i, column=1, value=insight)
            cell.font = Font(size=10)
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
    
    def add_legend_section(self, ws):
        """Add legend section with KPMG styling."""
        # Add Legend section header
        header_cell = ws.cell(row=1, column=16, value="Legend")
        header_cell.font = Font(size=12, bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color=self.kpmg_dark_blue, end_color=self.kpmg_dark_blue, fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('P1:R1')
        
        # Add sub-headers
        ws['P2'] = "Rating"
        ws['Q2'] = "CVSS/VPR Score"
        ws['R2'] = "Priority Score"
        
        for col in range(16, 19):
            cell = ws.cell(row=2, column=col)
            style_header_cell(cell, bg_color=self.kpmg_light_blue)
        
        # Add CVSS rating data with styling
        ratings = [
            ["None", "0", "0"],
            ["Low", "0.1 - 3.9", "1"],
            ["Medium", "4.0 - 6.9", "2"],
            ["High", "7.0 - 8.9", "3"],
            ["Critical", "9.0 - 10.0", "4"]
        ]
        
        # Add border for the entire table
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for i, (rating, score, priority) in enumerate(ratings, 3):
            for col in range(16, 19):
                ws.cell(row=i, column=col).border = border
            
            # Set values
            ws.cell(row=i, column=16, value=rating)
            ws.cell(row=i, column=17, value=score)
            ws.cell(row=i, column=18, value=int(priority))
            
            # Color rating cells according to risk level
            if rating in self.risk_colors:
                ws.cell(row=i, column=16).fill = PatternFill(
                    start_color=self.risk_colors[rating],
                    end_color=self.risk_colors[rating],
                    fill_type="solid"
                )
        
        # Add EPSS Score legend header
        header_cell = ws.cell(row=8, column=16, value="EPSS Score Legend")
        header_cell.font = Font(size=12, bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color=self.kpmg_dark_blue, end_color=self.kpmg_dark_blue, fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('P8:R8')
        
        # Add EPSS sub-headers
        ws['P9'] = "Rating"
        ws['Q9'] = "EPSS Score"
        ws['R9'] = "Priority Score"
        
        for col in range(16, 19):
            cell = ws.cell(row=9, column=col)
            style_header_cell(cell, bg_color=self.kpmg_light_blue)
        
        # Add EPSS ratings with styling
        epss_ratings = [
            ["None", "<0.01", "0"],
            ["Low", "0.01-0.39", "1"],
            ["Medium", "0.4-0.69", "2"],
            ["High", "0.7-0.89", "3"],
            ["Critical", "0.9-1", "4"]
        ]
        
        for i, (rating, score, priority) in enumerate(epss_ratings, 10):
            for col in range(16, 19):
                ws.cell(row=i, column=col).border = border
            
            # Set values
            ws.cell(row=i, column=16, value=rating)
            ws.cell(row=i, column=17, value=score)
            ws.cell(row=i, column=18, value=int(priority))
            
            # Color rating cells according to risk level
            if rating in self.risk_colors:
                ws.cell(row=i, column=16).fill = PatternFill(
                    start_color=self.risk_colors[rating],
                    end_color=self.risk_colors[rating],
                    fill_type="solid"
                )
        
        # Add KEV legend header
        header_cell = ws.cell(row=15, column=16, value="KEV Legend")
        header_cell.font = Font(size=12, bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color=self.kpmg_dark_blue, end_color=self.kpmg_dark_blue, fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('P15:R15')
        
        # Add KEV sub-headers
        ws['P16'] = "Status"
        ws['Q16'] = "Meaning"
        ws['R16'] = "Priority Score"
        
        for col in range(16, 19):
            cell = ws.cell(row=16, column=col)
            style_header_cell(cell, bg_color=self.kpmg_light_blue)
        
        # Add KEV ratings with styling
        kev_ratings = [
            ["Yes", "Listed in CISA KEV", "4"],
            ["No", "Not in CISA KEV", "0"]
        ]
        
        for i, (status, meaning, priority) in enumerate(kev_ratings, 17):
            for col in range(16, 19):
                ws.cell(row=i, column=col).border = border
            
            # Set values
            ws.cell(row=i, column=16, value=status)
            ws.cell(row=i, column=17, value=meaning)
            ws.cell(row=i, column=18, value=int(priority))
            
            # Color Yes with Critical color
            if status == "Yes":
                ws.cell(row=i, column=16).fill = PatternFill(
                    start_color=self.risk_colors["Critical"],
                    end_color=self.risk_colors["Critical"],
                    fill_type="solid"
                )
        
        # Add common priority formula section
        header_cell = ws.cell(row=20, column=16, value="Prioritization Formula")
        header_cell.font = Font(size=12, bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color=self.kpmg_dark_blue, end_color=self.kpmg_dark_blue, fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('P20:R20')
        
        # Add priority formula
        ws['P21'] = "Formula"
        ws['Q21'] = "EPSS + VPR + KEV"
        ws.merge_cells('Q21:R21')
        
        ws['P22'] = "Highest"
        ws['Q22'] = "12"
        ws.merge_cells('Q22:R22')
        
        ws['P23'] = "Lowest"
        ws['Q23'] = "1"
        ws.merge_cells('Q23:R23')
        
        # Add borders
        for row in range(21, 24):
            for col in range(16, 19):
                ws.cell(row=row, column=col).border = border
            
            cell = ws.cell(row=row, column=16)
            cell.fill = PatternFill(start_color=self.kpmg_light_blue, end_color=self.kpmg_light_blue, fill_type="solid")
    
    def add_summary_bar_chart(self, ws):
        """Add a summary bar chart for risk counts at cell A17."""
        from openpyxl.chart import BarChart3D, Reference
        # Data range: B2:F3 (headers and counts)
        data = Reference(ws, min_col=9, min_row=20, max_col=12, max_row=24)
        categories = Reference(ws, min_col=8, min_row=21, max_col=8, max_row=24)
        chart = BarChart3D()
        chart.type = "col"
        chart.style = 10        
        chart.title = "Risk Count Summary"
        chart.y_axis.title = 'Count'
        chart.x_axis.title = 'Metric'
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 10
        chart.width = 18
        ws.add_chart(chart, "A17")
    
    def get_common_priority_vulns(self, processed_df):
        """Return vulnerabilities that are High/Critical in at least 2 of EPSS, VPR, KEV."""
        if processed_df is None:
            return []
        df = processed_df.copy()
        # Map KEV to High if Yes, else Low
        kev_cat = df['KEV Listed'].map(lambda x: 'High' if x == 'Yes' else 'Low')
        # Count how many of EPSS, VPR, KEV are High or Critical
        high_crit = df[['EPSS Category', 'VPR Category']].apply(lambda x: [(v in ['High', 'Critical']) for v in x], axis=1)
        kev_high = kev_cat.isin(['High', 'Critical'])
        count = high_crit.apply(sum) + kev_high.astype(int)
        mask = count >= 2
        result = df[mask]
        return result

    def add_common_priority_summary(self, ws, processed_df, start_row=27):
        """Add a mini table summarizing common high/critical vulnerabilities (at least 2 of EPSS, VPR, KEV)."""
        import pandas as pd
        common_df = self.get_common_priority_vulns(processed_df)
        # Ensure common_df is a DataFrame
        if isinstance(common_df, list):
            common_df = pd.DataFrame(common_df)
        count = len(common_df)
        ws.cell(row=start_row, column=8, value="Common Priority Vulnerabilities (High/Critical in ≥2 of EPSS, VPR, KEV)")
        ws.merge_cells(start_row=start_row, start_column=8, end_row=start_row, end_column=12)
        ws.cell(row=start_row, column=8).font = Font(size=11, bold=True, color=self.kpmg_dark_blue)
        ws.cell(row=start_row+1, column=8, value="Count")
        ws.cell(row=start_row+1, column=9, value=count)
        ws.cell(row=start_row+2, column=8, value="Example CVEs")
        # Show up to 3 example CVEs
        cves = list(common_df['CVE'].dropna().unique())[:3] if 'CVE' in common_df.columns else []
        ws.cell(row=start_row+2, column=9, value=", ".join(cves) if cves else "-")
        # Style
        for r in range(start_row, start_row+3):
            for c in range(8, 10):
                ws.cell(row=r, column=c).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                ws.cell(row=r, column=c).fill = PatternFill(start_color=self.kpmg_light_blue, end_color=self.kpmg_light_blue, fill_type="solid")