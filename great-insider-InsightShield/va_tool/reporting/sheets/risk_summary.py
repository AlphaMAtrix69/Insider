"""Risk Summary sheet generator."""

import pandas as pd
import matplotlib.pyplot as plt
import os
from openpyxl.drawing.image import Image

from va_tool.reporting.sheets.base import BaseSheetGenerator
from va_tool.utils import ensure_dir_exists


class RiskSummarySheetGenerator(BaseSheetGenerator):
    """Generator for the Risk Summary sheet."""
    
    def __init__(self):
        """Initialize the generator."""
        super().__init__(title="5.1 Risk Summary")
    
    def generate(self, wb, df=None, output_dir=None, **kwargs):
        """
        Generate the Risk Summary sheet.
        
        Args:
            wb: Excel workbook
            df: DataFrame with vulnerability data
            output_dir: Directory to save chart images
            **kwargs: Additional arguments
        
        Returns:
            The worksheet
        """
        self.logger.info("Generating Risk Summary sheet")
        ws = super().generate(wb)
        
        # Add title
        self.add_title(
            ws, "Vulnerability Risk Summary", 
            font_size=14, merge_range='A1:C1'
        )
        
        # Define severity order
        severity_order = ['Critical', 'High', 'Medium', 'Low/None']
        
        # Process data
        if df is not None and 'Risk' in df.columns:
            # Count vulnerabilities by severity, combining Low and None
            risk_counts = df['Risk'].replace('None', 'Low/None').replace('Low', 'Low/None').value_counts().reset_index()
            risk_counts.columns = ['Severity', 'Count']
            
            # Ensure all severities are included, even if count is 0
            risk_summary = pd.DataFrame({'Severity': severity_order})
            risk_summary = risk_summary.merge(risk_counts, on='Severity', how='left').fillna({'Count': 0})
            risk_summary['Count'] = risk_summary['Count'].astype(int)
            
            # Add column for vulnerabilities with CVE
            if 'CVE' in df.columns:
                risk_summary['Vulnerabilities with CVE'] = risk_summary['Severity'].apply(
                    lambda x: df[df['Risk'].isin(['Low', 'None'])]['CVE'].apply(lambda y: pd.notna(y) and y != '').sum()
                    if x == 'Low/None' else
                    df[df['Risk'] == x]['CVE'].apply(lambda y: pd.notna(y) and y != '').sum()
                )
            else:
                risk_summary['Vulnerabilities with CVE'] = 0
            
            # Set column widths
            ws.column_dimensions['A'].width = 20  # Severity
            ws.column_dimensions['B'].width = 10  # Count
            ws.column_dimensions['C'].width = 25  # Vulnerabilities with CVE
            
            # Write headers
            headers = ['Severity', 'Count', 'Vulnerabilities with CVE']
            self.write_headers(ws, headers, row=3)
            
            # Write data
            for row_idx, row in enumerate(risk_summary.itertuples(), 4):
                ws.cell(row=row_idx, column=1, value=row[1])  # Severity
                ws.cell(row=row_idx, column=2, value=row[2])  # Count
                ws.cell(row=row_idx, column=3, value=row[3])  # Vulnerabilities with CVE
            
            # Generate pie chart if output_dir is provided
            if output_dir:
                self.generate_pie_chart(risk_summary, output_dir, ws)
        else:
            ws['A3'] = "No vulnerability data or 'Risk' column available."
            ws.column_dimensions['A'].width = 40
        
        return ws
    
    def generate_pie_chart(self, risk_summary, output_dir, ws):
        """Generate a pie chart for risk summary."""
        plt.figure(figsize=(8, 6))
        colors = ['#d62728', '#ff7f0e', '#ffbb78', '#2ca02c']  # Critical, High, Medium, Low/None
        
        # Create pie chart
        plt.pie(
            risk_summary['Count'],
            labels=risk_summary['Severity'],
            colors=colors,
            autopct='%1.1f%%',
            startangle=90,
            textprops={'fontsize': 12}
        )
        
        plt.title('Vulnerability Count by Severity', fontsize=14)
        plt.axis('equal')  # Ensure pie is circular
        plt.tight_layout()
        
        # Save chart
        ensure_dir_exists(output_dir)
        chart_path = os.path.join(output_dir, 'risk_chart.png')
        plt.savefig(chart_path, bbox_inches='tight', facecolor='white')
        plt.close()
        
        # Add chart image to Excel
        try:
            img = Image(chart_path)
            img.width = 600
            img.height = 400
            ws.add_image(img, 'A10')
            self.logger.info(f"Added risk chart to sheet: {chart_path}")
        except Exception as e:
            self.logger.error(f"Error adding risk chart: {str(e)}")
            ws['A10'] = "Error adding chart image"