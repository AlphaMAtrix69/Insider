"""Excel styling utilities for vulnerability analysis reports."""

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def style_header_cell(cell, bold=True, bg_color="D3D3D3"):
    """Style header cells with uniform look."""
    cell.font = Font(bold=bold)
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border


def write_df_to_sheet(ws, df, start_row=1, start_col=1, include_header=True, style_header=True):
    """Write a DataFrame to worksheet with optional styling."""
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    rows = dataframe_to_rows(df, index=False, header=include_header)
    
    for r_idx, row in enumerate(rows, start_row):
        for c_idx, value in enumerate(row, start_col):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Style header row
            if r_idx == start_row and include_header and style_header:
                style_header_cell(cell)


def set_column_widths(ws, columns, width=15):
    """Set column widths for a worksheet."""
    from openpyxl.utils import get_column_letter
    
    for col in range(1, columns + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = width