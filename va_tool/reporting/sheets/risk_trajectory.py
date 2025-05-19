"""Risk Trajectory sheet generator."""

import pandas as pd
import matplotlib.pyplot as plt
import os
import re
from datetime import datetime
from openpyxl.drawing.image import Image

from va_tool.reporting.sheets.base import BaseSheetGenerator
from va_tool.utils import ensure_dir_exists, style_header_cell


class RiskTrajectorySheetGenerator(BaseSheetGenerator):
    """Generator for the Risk Trajectory sheet."""
    
    def __init__(self):
        """Initialize the generator."""
        super().__init__(title="5.4 Risk Trajectory")
    
    def generate(self, wb, df=None, output_dir=None, **kwargs):
        """
        Generate the Risk Trajectory sheet.
        
        Args:
            wb: Excel workbook
            df: DataFrame with vulnerability data
            output_dir: Directory to save chart images
            **kwargs: Additional arguments
        
        Returns:
            The worksheet
        """
        self.logger.info("Generating Risk Trajectory sheet")
        ws = super().generate(wb)
        
        # Add titles
        self.add_title(
            ws, "Risk Trajectory Security Posture Shift", 
            font_size=14, merge_range='A1:D1'
        )
        
        self.add_title(
            ws, "Vulnerability Counts by Severity", 
            cell="F1", font_size=14, merge_range='F1:J1'
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 20  # Timestamp
        ws.column_dimensions['B'].width = 25  # Total Vulnerabilities
        ws.column_dimensions['C'].width = 25  # Vulnerabilities with CVEs
        ws.column_dimensions['D'].width = 30  # % Change (left)
        ws.column_dimensions['F'].width = 20  # Critical
        ws.column_dimensions['G'].width = 20  # High
        ws.column_dimensions['H'].width = 20  # Medium
        ws.column_dimensions['I'].width = 20  # Low/None
        ws.column_dimensions['J'].width = 40  # % Change (right)
        
        # Get historical data if output_dir is provided
        if output_dir:
            # Current timestamp
            current_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            current_dt = datetime.now()
            
            # Get historical data
            try:
                data = self.get_historical_data(output_dir, current_dt)
                
                # Add current data
                current_data = self.get_current_data(df, current_timestamp, current_dt)
                if current_data:
                    data.append(current_data)
                
                # Sort by datetime
                data.sort(key=lambda x: x[1])
                
                # Calculate percent changes
                percent_changes = self.calculate_percent_changes(data)
                
                # Create DataFrames for displaying in sheet
                traj_summary, severity_summary = self.create_summary_dataframes(data, percent_changes)
                
                # Write tables
                if not traj_summary.empty:
                    # Write left table headers
                    headers_left = ['Timestamp', 'Total Vulnerabilities', 'Vulnerabilities with CVEs', '% Change']
                    self.write_headers(ws, headers_left, row=3)
                    
                    # Write left table data
                    for row_idx, row in enumerate(traj_summary.itertuples(), 4):
                        ws.cell(row=row_idx, column=1, value=row[1])  # Timestamp
                        ws.cell(row=row_idx, column=2, value=row[2])  # Total Vulnerabilities
                        ws.cell(row=row_idx, column=3, value=row[3])  # Vulnerabilities with CVEs
                        ws.cell(row=row_idx, column=4, value=row[4])  # % Change
                    
                    # Write right table headers
                    headers_right = ['Critical', 'High', 'Medium', 'Low/None', '% Change']
                    for col_idx, header in enumerate(headers_right, 6):  # Start at F (6)
                        cell = ws.cell(row=3, column=col_idx, value=header)
                        style_header_cell(cell)
                    
                    # Write right table data
                    for row_idx, row in enumerate(severity_summary.itertuples(), 4):
                        ws.cell(row=row_idx, column=6, value=row[1])  # Critical
                        ws.cell(row=row_idx, column=7, value=row[2])  # High
                        ws.cell(row=row_idx, column=8, value=row[3])  # Medium
                        ws.cell(row=row_idx, column=9, value=row[4])  # Low/None
                        ws.cell(row=row_idx, column=10, value=row[5])  # % Change
                    
                    # Generate charts
                    self.generate_trajectory_charts(traj_summary, severity_summary, output_dir, ws)
                else:
                    ws['A3'] = "No historical vulnerability data available."
            except Exception as e:
                self.logger.error(f"Error generating risk trajectory: {str(e)}")
                ws['A3'] = f"Error generating risk trajectory: {str(e)}"
        else:
            ws['A3'] = "Output directory not provided for historical data analysis."
        
        return ws
    
    def get_historical_data(self, output_dir, current_dt):
        """Get historical vulnerability data from previous reports."""
        data = []
        pattern = r'Enhanced_Vulnerability_Analysis_(\d{8}_\d{6})\.xlsx'
        
        try:
            files = [f for f in os.listdir(output_dir) if re.match(pattern, f)]
            for file in files:
                match = re.match(pattern, file)
                if match:
                    timestamp = match.group(1)  # YYYYMMDD_HHMMSS
                    try:
                        file_dt = datetime.strptime(timestamp, "%Y%m%d_%H%M%S")
                        file_path = os.path.join(output_dir, file)
                        
                        # Read data from previous report
                        try:
                            # Load 5.1 Risk Summary sheet
                            df = pd.read_excel(file_path, sheet_name="5.1 Risk Summary", skiprows=2)
                            
                            # Sum 'Count' for total vulnerabilities
                            total_vulns = df['Count'].sum()
                            
                            # Sum 'Vulnerabilities with CVE' (if exists)
                            cve_vulns = df[
                                'Vulnerabilities with CVE'].sum() if 'Vulnerabilities with CVE' in df.columns else 0
                            
                            # Get severity counts
                            severity_counts = {'Critical': 0, 'High': 0, 'Medium': 0, 'Low/None': 0}
                            for _, row in df.iterrows():
                                severity = row['Severity']
                                if severity in severity_counts or severity in ['Low', 'None']:
                                    severity = 'Low/None' if severity in ['Low', 'None'] else severity
                                    severity_counts[severity] = row['Count']
                            
                            # Add to data list
                            data.append((
                                timestamp, file_dt, total_vulns, cve_vulns,
                                severity_counts['Critical'], severity_counts['High'],
                                severity_counts['Medium'], severity_counts['Low/None']
                            ))
                        except Exception as e:
                            self.logger.warning(f"Could not process {file}: {e}")
                    except ValueError:
                        self.logger.warning(f"Invalid timestamp in {file}")
        except FileNotFoundError:
            self.logger.warning(f"Output directory {output_dir} not found.")
        
        return data
    
    def get_current_data(self, df, current_timestamp, current_dt):
        """Get current vulnerability data."""
        if df is None:
            return None
        
        # Calculate total vulnerabilities
        current_total_vulns = len(df)
        
        # Calculate vulnerabilities with CVEs
        current_cve_vulns = 0
        if 'CVE' in df.columns:
            current_cve_vulns = df['CVE'].apply(lambda x: pd.notna(x) and x != '').sum()
        
        # Calculate severity counts
        current_severity_counts = {'Critical': 0, 'High': 0, 'Medium': 0, 'Low/None': 0}
        if 'Risk' in df.columns:
            severity_counts = df['Risk'].replace('None', 'Low/None').replace('Low', 'Low/None').value_counts()
            for severity in current_severity_counts:
                current_severity_counts[severity] = severity_counts.get(severity, 0)
        
        # Return current data
        return (
            current_timestamp, current_dt, current_total_vulns, current_cve_vulns,
            current_severity_counts['Critical'], current_severity_counts['High'],
            current_severity_counts['Medium'], current_severity_counts['Low/None']
        )
    
    def calculate_percent_changes(self, data):
        """Calculate percent changes between consecutive data points."""
        percent_changes_left = []
        percent_changes_right = []
        
        for i in range(len(data)):
            if i == 0:  # First run: no previous data
                percent_changes_left.append("N/A")
                percent_changes_right.append("N/A")
            else:
                prev = data[i - 1]
                curr = data[i]
                
                # Left table: Total Vulnerabilities, Vulnerabilities with CVEs
                total_change = self.calculate_percent_change(curr[2], prev[2])
                cve_change = self.calculate_percent_change(curr[3], prev[3])
                left_change = f"Total: {total_change}, CVEs: {cve_change}"
                
                # Right table: Critical, High, Medium, Low/None
                crit_change = self.calculate_percent_change(curr[4], prev[4])
                high_change = self.calculate_percent_change(curr[5], prev[5])
                med_change = self.calculate_percent_change(curr[6], prev[6])
                low_change = self.calculate_percent_change(curr[7], prev[7])
                right_change = f"Crit: {crit_change}, High: {high_change}, Med: {med_change}, Low: {low_change}"
                
                percent_changes_left.append(left_change)
                percent_changes_right.append(right_change)
        
        return (percent_changes_left, percent_changes_right)
    
    def calculate_percent_change(self, current, previous):
        """Calculate percent change between two values."""
        if previous is None or previous == 0:
            return "N/A"
        try:
            change = ((current - previous) / previous) * 100
            if change >= 0:
                return f"+{int(round(change))}%"
            else:
                return f"{int(round(change))}%"
        except (TypeError, ValueError):
            return "N/A"
    
    def format_time_diff(self, current_dt, past_dt):
        """Format time difference for display."""
        if current_dt.date() == past_dt.date():
            return past_dt.strftime("%H:%M")
        elif current_dt.year == past_dt.year and current_dt.month == past_dt.month:
            return past_dt.strftime("%d/%m")
        elif current_dt.year == past_dt.year:
            return past_dt.strftime("%m/%y")
        else:
            return past_dt.strftime("%Y")
    
    def create_summary_dataframes(self, data, percent_changes):
        """Create summary DataFrames for display in the sheet."""
        if not data:
            return pd.DataFrame(), pd.DataFrame()
        
        # Unpack percent changes
        percent_changes_left, percent_changes_right = percent_changes
        
        # Format timestamps
        current_dt = data[-1][1]  # Use the latest timestamp as current
        timestamps = []
        for _, dt, _, _, _, _, _, _ in data:
            if dt == current_dt:
                timestamps.append("Now")
            else:
                timestamps.append(self.format_time_diff(current_dt, dt))
        
        # Create trajectory summary DataFrame
        traj_summary = pd.DataFrame({
            'Timestamp': timestamps,
            'Total Vulnerabilities': [x[2] for x in data],
            'Vulnerabilities with CVEs': [x[3] for x in data],
            '% Change': percent_changes_left
        })
        
        # Create severity summary DataFrame
        severity_summary = pd.DataFrame({
            'Critical': [x[4] for x in data],
            'High': [x[5] for x in data],
            'Medium': [x[6] for x in data],
            'Low/None': [x[7] for x in data],
            '% Change': percent_changes_right
        })
        
        return traj_summary, severity_summary
    
    def generate_trajectory_charts(self, traj_summary, severity_summary, output_dir, ws):
        """Generate trajectory charts for vulnerabilities over time."""
        # Generate left line chart (total vulnerabilities)
        if len(traj_summary) > 0:
            plt.figure(figsize=(8, 4))
            plt.plot(range(len(traj_summary)), traj_summary['Total Vulnerabilities'], 
                    marker='o', color='#1f77b4', label='Total Vulnerabilities')
            
            plt.xlabel('Timestamp', fontsize=12)
            plt.ylabel('Total Vulnerabilities', fontsize=12)
            plt.title('Total Vulnerabilities Trend Over Time', fontsize=14)
            plt.xticks(range(len(traj_summary)), traj_summary['Timestamp'], rotation=45)
            plt.grid(True)
            
            # Add value labels
            for i, count in enumerate(traj_summary['Total Vulnerabilities']):
                plt.text(i, count + 0.5, str(count), ha='center', fontsize=10)
            
            plt.legend()
            plt.tight_layout()
            
            # Save chart
            ensure_dir_exists(output_dir)
            chart_path = os.path.join(output_dir, 'risk_trajectory_chart.png')
            plt.savefig(chart_path, bbox_inches='tight', facecolor='white')
            plt.close()
            
            # Add chart to worksheet
            last_table_row = 3 + len(traj_summary)
            chart_row = last_table_row + 3
            try:
                img = Image(chart_path)
                img.width = 600
                img.height = 300
                ws.add_image(img, f'A{chart_row}')
            except Exception as e:
                self.logger.error(f"Error adding trajectory chart: {str(e)}")
                ws[f'A{chart_row}'] = f"Error adding chart: {str(e)}"
        
        # Generate right line chart (severity trends)
        if len(severity_summary) > 0:
            plt.figure(figsize=(8, 4))
            colors = {'Critical': '#d62728', 'High': '#ff7f0e', 'Medium': '#ffbb78', 'Low/None': '#2ca02c'}
            
            for severity, color in colors.items():
                plt.plot(range(len(severity_summary)), severity_summary[severity], 
                        marker='o', color=color, label=severity)
            
            plt.xlabel('Timestamp', fontsize=12)
            plt.ylabel('Vulnerability Count', fontsize=12)
            plt.title('Vulnerability Trends by Severity', fontsize=14)
            plt.xticks(range(len(severity_summary)), traj_summary['Timestamp'], rotation=45)
            plt.grid(True)
            
            # Add value labels
            for severity in colors:
                for i, count in enumerate(severity_summary[severity]):
                    if count > 0:  # Annotate only non-zero counts
                        plt.text(i, count + 0.5, str(count), ha='center', fontsize=8)
            
            plt.legend()
            plt.tight_layout()
            
            # Save chart
            severity_chart_path = os.path.join(output_dir, 'severity_trajectory_chart.png')
            plt.savefig(severity_chart_path, bbox_inches='tight', facecolor='white')
            plt.close()
            
            # Add chart to worksheet
            try:
                img = Image(severity_chart_path)
                img.width = 600
                img.height = 300
                ws.add_image(img, f'F{chart_row}')
            except Exception as e:
                self.logger.error(f"Error adding severity trajectory chart: {str(e)}")
                ws[f'F{chart_row}'] = f"Error adding chart: {str(e)}"