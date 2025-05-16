"""Vulnerability Density sheet generator."""

import pandas as pd
import matplotlib.pyplot as plt
import os
from openpyxl.drawing.image import Image

from va_tool.reporting.sheets.base import BaseSheetGenerator
from va_tool.utils import ensure_dir_exists


class VulnDensitySheetGenerator(BaseSheetGenerator):
    """Generator for the Top 10 Vulnerable Hosts sheet."""
    
    def __init__(self):
        """Initialize the generator."""
        super().__init__(title="5.3 Top 10 Vulnerable Hosts")
    
    def generate(self, wb, df=None, output_dir=None, **kwargs):
        """
        Generate the Top 10 Vulnerable Hosts sheet.
        
        Args:
            wb: Excel workbook
            df: DataFrame with vulnerability data
            output_dir: Directory to save chart images
            **kwargs: Additional arguments
        
        Returns:
            The worksheet
        """
        self.logger.info("Generating Top 10 Vulnerable Hosts sheet")
        ws = super().generate(wb)
        
        # Add title
        self.add_title(
            ws, "Top 10 Vulnerable Hosts", 
            font_size=14, merge_range='A1:G1'
        )
        
        # Process data
        if df is not None and 'Host' in df.columns and 'Risk' in df.columns:
            # Prepare vulnerability summary by host
            vuln_summary = self.prepare_host_summary(df)
            
            # Set column widths
            ws.column_dimensions['A'].width = 30  # Host
            for col in 'BCDEFG':
                ws.column_dimensions[col].width = 20  # Other columns
            
            # Write headers
            headers = ['Host', 'Vulnerability Count', 'Vulnerabilities with CVE', 
                      'Critical', 'High', 'Medium', 'Low/None']
            self.write_headers(ws, headers, row=3)
            
            # Write data
            for row_idx, row in enumerate(vuln_summary.itertuples(), 4):
                ws.cell(row=row_idx, column=1, value=row[1])  # Host
                ws.cell(row=row_idx, column=2, value=row[2])  # Vulnerability Count
                ws.cell(row=row_idx, column=3, value=row[3])  # Vulnerabilities with CVE
                ws.cell(row=row_idx, column=4, value=row[4])  # Critical
                ws.cell(row=row_idx, column=5, value=row[5])  # High
                ws.cell(row=row_idx, column=6, value=row[6])  # Medium
                ws.cell(row=row_idx, column=7, value=row[7])  # Low/None
            
            # Generate stacked horizontal bar chart if output_dir is provided
            if output_dir:
                self.generate_stacked_bar_chart(vuln_summary, output_dir, ws)
        else:
            ws['A3'] = "No vulnerability data or required columns available."
            ws.column_dimensions['A'].width = 40
        
        return ws
    
    def prepare_host_summary(self, df):
        """Prepare summary of vulnerabilities by host."""
        # Count total vulnerabilities by host
        vuln_counts = df['Host'].value_counts().reset_index()
        vuln_counts.columns = ['Host', 'Vulnerability Count']
        
        # Calculate vulnerabilities with CVE
        if 'CVE' in df.columns:
            cve_counts = df[df['CVE'].apply(lambda x: pd.notna(x) and x != '')]['Host'].value_counts().reset_index()
            cve_counts.columns = ['Host', 'Vulnerabilities with CVE']
            vuln_summary = vuln_counts.merge(cve_counts, on='Host', how='left').fillna({'Vulnerabilities with CVE': 0})
            vuln_summary['Vulnerabilities with CVE'] = vuln_summary['Vulnerabilities with CVE'].astype(int)
        else:
            vuln_summary = vuln_counts.copy()
            vuln_summary['Vulnerabilities with CVE'] = 0
        
        # Calculate severity counts
        severity_counts = df.copy()
        severity_counts['Risk'] = severity_counts['Risk'].replace('None', 'Low/None').replace('Low', 'Low/None')
        
        severity_pivot = severity_counts.pivot_table(
            index='Host',
            columns='Risk',
            aggfunc='size',
            fill_value=0
        ).reset_index()
        
        # Ensure all severity levels exist in the dataframe
        severity_cols = ['Critical', 'High', 'Medium', 'Low/None']
        for col in severity_cols:
            if col not in severity_pivot.columns:
                severity_pivot[col] = 0
        
        # Select relevant columns
        severity_pivot = severity_pivot[['Host'] + severity_cols]
        
        # Merge all counts
        vuln_summary = vuln_summary.merge(severity_pivot, on='Host', how='left').fillna(0)
        
        # Convert numeric columns to integers
        for col in severity_cols:
            vuln_summary[col] = vuln_summary[col].astype(int)
        
        # Sort by vulnerability count and limit to top 10
        vuln_summary = vuln_summary.sort_values('Vulnerability Count', ascending=False).head(10)
        
        return vuln_summary
    
    def generate_stacked_bar_chart(self, vuln_summary, output_dir, ws):
        """Generate a stacked horizontal bar chart for vulnerability distribution."""
        try:
            plt.figure(figsize=(12, 6))
            
            # Define severity columns and colors
            severity_cols = ['Critical', 'High', 'Medium', 'Low/None']
            colors = ['#d62728', '#ff7f0e', '#ffbb78', '#2ca02c']  # Red, Orange, Light Orange, Green
            
            # Create stacked horizontal bar chart
            left = pd.Series(0, index=vuln_summary.index)
            
            for col, color in zip(severity_cols, colors):
                plt.barh(vuln_summary['Host'], vuln_summary[col], left=left, color=color, label=col)
                left += vuln_summary[col]
            
            # Add labels and formatting
            plt.xlabel('Vulnerability Count', fontsize=12)
            plt.title('Top 10 Hosts by Vulnerability Count (Severity Breakdown)', fontsize=14)
            plt.gca().invert_yaxis()  # Highest count at top
            plt.legend(title='Severity')
            
            # Add total count labels
            for i, total in enumerate(vuln_summary['Vulnerability Count']):
                plt.text(total, i, str(total), va='center', fontsize=10)
            
            plt.tight_layout()
            
            # Save chart
            ensure_dir_exists(output_dir)
            chart_path = os.path.join(output_dir, 'vuln_density_chart.png')
            plt.savefig(chart_path, bbox_inches='tight', facecolor='white')
            plt.close()
            
            # Add chart to Excel
            img = Image(chart_path)
            img.width = 950
            img.height = 400
            ws.add_image(img, 'A15')
            self.logger.info(f"Added vulnerability density chart: {chart_path}")
        
        except Exception as e:
            self.logger.error(f"Error generating vulnerability density chart: {str(e)}")
            ws['A15'] = f"Error generating chart: {str(e)}"