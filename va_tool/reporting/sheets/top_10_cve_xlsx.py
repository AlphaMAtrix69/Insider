from collections import Counter
from openpyxl.chart import BarChart3D, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class Top10CVEsSheetMixin:
    def generate_top_10_cves(self, ws, df, output_dir=None):
        """Generate Top 10 CVEs table and 3D bar chart using consistent styling."""
        try:
            if not {'CVE', 'Risk'}.issubset(df.columns):
                raise ValueError("DataFrame missing required columns: 'CVE' and/or 'Risk'")

            df_cves = df[["CVE", "Risk"]].dropna(subset=["CVE"])
            all_cves = []
            cve_severities = {}
            precedence = {'Critical': 4, 'High': 3, 'Medium': 2, 'Low': 1, 'None': 1, 'Low/None': 1}

            for _, row in df_cves.iterrows():
                cves = [cve.strip() for cve in str(row["CVE"]).split(',') if cve.strip().startswith('CVE')]
                severity = 'Low/None' if row["Risk"] in ['Low', 'None'] else row["Risk"]
                for cve in cves:
                    all_cves.append(cve)
                    if (cve not in cve_severities or
                        precedence.get(severity, 0) > precedence.get(cve_severities.get(cve, ""), 0)):
                        cve_severities[cve] = severity

            top_10 = Counter(all_cves).most_common(10)
            if not top_10:
                ws['J1'] = "No CVEs found"
                return

            # Add title with consistent styling
            self.add_title(ws, "Top 10 CVEs", font_size=14, merge_range='J1:L1', cell="J1", bold=True)


            # Headers
            headers = ["CVE", "Count", "Severity"]
            for col, header in enumerate(headers, start=10):  # Column J = 10
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin")
                )

            else:
                # Manual fallback
                for col, header in enumerate(headers, start=10):
                    cell = ws.cell(row=3, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin")
                    )

            # Data rows
            for i, (cve, count) in enumerate(top_10, start=4):
                ws.cell(row=i, column=10, value=cve).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=i, column=11, value=count).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=i, column=12, value=cve_severities.get(cve, "N/A")).alignment = Alignment(horizontal='center', vertical='center')

            # Set column widths
            ws.column_dimensions["J"].width = 20
            ws.column_dimensions["K"].width = 10
            ws.column_dimensions["L"].width = 15

            # Chart
            chart = BarChart3D()
            chart.title = "Top 10 CVEs by Count"
            chart.style = 10
            chart.y_axis.title = "Occurrences"
            chart.x_axis.title = "CVE"
            chart.height = 10
            chart.width = 20
            chart.legend.overlay = False
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.dataLabels.showCatName = False
            chart.dataLabels.showPercent = False
            chart.dataLabels.showSerName = False

            data = Reference(ws, min_col=11, min_row=4, max_row=3 + len(top_10))
            cats = Reference(ws, min_col=10, min_row=4, max_row=3 + len(top_10))
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)

            ws.add_chart(chart, "J16")

        except Exception as e:
            ws['J1'] = "Error generating Top 10 CVEs"
            ws['J2'] = str(e)
