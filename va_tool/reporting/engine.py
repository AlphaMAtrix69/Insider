"""Report engine that coordinates all report generation."""

import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from va_tool.utils import get_logger, format_datetime, ensure_dir_exists
from va_tool.reporting.excel_writer import save_excel_workbook, write_full_excel_report
from va_tool.reporting.json_writer import write_json_report
from va_tool.reporting.sheets import (
    SummarySheetGenerator,
    PrioritizationInsightsGenerator,
    ExploitabilitySheetGenerator,
    AgeingSheetGenerator,
    MostExploitableSheetGenerator,
    RiskSummarySheetGenerator,
    CveSummarySheetGenerator,
    HostsSummarySheetGenerator,
    IPInsightsSheetGenerator,
    VulnClusteringSheetGenerator,
    UniqueVulnSheetGenerator,
    EOLSummarySheetGenerator,
    EOLComponentsSheetGenerator,
    EOLIPsSheetGenerator,
    EOLVersionsSheetGenerator
)

logger = get_logger()


class ReportEngine:
    """Coordinates the generation of reports and sheets."""
    
    def __init__(self, output_dir):
        """
        Initialize the report engine.
        
        Args:
            output_dir: Directory to save reports
        """
        self.output_dir = output_dir
        ensure_dir_exists(output_dir)
        
        # KPMG color scheme
        self.kpmg_blue = "00338D"
        self.kpmg_light_blue = "DCE6F1"
        self.kpmg_dark_blue = "005EB8"
        self.kpmg_medium_blue = "0091DA"
        self.kpmg_accent_blue = "00A3E0"
        
        # Initialize sheet generators
        self.sheet_generators = {
            'summary': SummarySheetGenerator(),
            'prioritization': PrioritizationInsightsGenerator(),
            'exploitability': ExploitabilitySheetGenerator(),
            'ageing': AgeingSheetGenerator(),
            'most_exploitable': MostExploitableSheetGenerator(),
            'risk_summary': RiskSummarySheetGenerator(),
            'top10_cves': CveSummarySheetGenerator(),
            'host_summary': HostsSummarySheetGenerator(),
            # 'risk_trajectory': RiskTrajectorySheetGenerator(),
            'ip_insights': IPInsightsSheetGenerator(),
            'vulnerability_clustering': VulnClusteringSheetGenerator(),
            'unique_vuln': UniqueVulnSheetGenerator(),
            'eol_summary': EOLSummarySheetGenerator(),
            'eol_components': EOLComponentsSheetGenerator(),
            'eol_ips': EOLIPsSheetGenerator(),
            'eol_versions': EOLVersionsSheetGenerator()
        }
    
    def create_standard_report(self, original_df, processed_df, check_needed_df, analyzed_df, results_data):
        """
        Create a standard Excel report with multiple sheets.
        
        Args:
            original_df: Original vulnerability DataFrame
            processed_df: Processed vulnerability DataFrame
            check_needed_df: Vulnerabilities that need checking
            analyzed_df: Analyzed vulnerability DataFrame
            results_data: Dictionary with analysis results
        
        Returns:
            Path to the saved Excel file
        """
        logger.info("Creating standard Excel report")
        return write_full_excel_report(
            original_df, processed_df, check_needed_df, analyzed_df, 
            results_data, self.output_dir, "Vulnerability_Analysis"
        )
    
    def create_json_report(self, results_data):
        """
        Create a JSON report with analysis results.
        
        Args:
            results_data: Dictionary with analysis results
        
        Returns:
            Path to the saved JSON file
        """
        logger.info("Creating JSON report")
        return write_json_report(results_data, self.output_dir)
    
    def create_enhanced_report(self, original_df, processed_df, analyzed_df, results_data, excluded_df=None):
        """
        Create an enhanced Excel report with visualizations.
        
        Args:
            original_df: Original vulnerability DataFrame
            processed_df: Processed vulnerability DataFrame
            analyzed_df: Analyzed vulnerability DataFrame
            results_data: Dictionary with analysis results
            excluded_df: DataFrame with Informational/None/Check Needed
        
        Returns:
            Path to the saved Excel file
        """
        logger.info("Creating enhanced Excel report with KPMG styling")
        
        # Create new workbook
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Apply workbook properties
        wb.properties.creator = "Vulnerability Analysis Tool"
        wb.properties.title = "KPMG Vulnerability Assessment Report"
        wb.properties.subject = "Security Vulnerability Analysis"
        wb.properties.created = datetime.datetime.now()
        
        # Generate sheets in improved logical order
        # Section 1: Prioritization and related
        self.sheet_generators['summary'].generate(wb, results_data, original_df, processed_df)
        self.sheet_generators['prioritization'].generate(wb, results_data, original_df, processed_df, excluded_df=excluded_df)
        self.sheet_generators['exploitability'].generate(wb, processed_df)
        self.sheet_generators['ageing'].generate(wb, analyzed_df)
        self.sheet_generators['most_exploitable'].generate(wb, processed_df)
        # Section 3: Operational/Insights
        self.sheet_generators['ip_insights'].generate(wb, processed_df, self.output_dir)
        self.sheet_generators['vulnerability_clustering'].generate(wb, processed_df, self.output_dir)
        self.sheet_generators['unique_vuln'].generate(wb, processed_df, self.output_dir)
        # Insert 3.4 Bucket Details after 3.3 Unique Vulnerabilities
        from va_tool.reporting.sheets.bucket_details import BucketDetailsSheetGenerator
        BucketDetailsSheetGenerator().generate(wb, processed_df)
        # Section 4: Management/Executive
        self.sheet_generators['risk_summary'].generate(wb, processed_df, self.output_dir)
        self.sheet_generators['top10_cves'].generate(wb, processed_df, self.output_dir)
        self.sheet_generators['host_summary'].generate(wb, processed_df, self.output_dir)
        # Section 5: EOL
        self.sheet_generators['eol_summary'].generate(wb, processed_df, self.output_dir, results_data)
        self.sheet_generators['eol_components'].generate(wb, processed_df, self.output_dir, results_data)
        self.sheet_generators['eol_ips'].generate(wb, processed_df, self.output_dir, results_data)
        self.sheet_generators['eol_versions'].generate(wb, processed_df, self.output_dir, results_data)
        # self.sheet_generators['risk_trajectory'].generate(wb, processed_df, self.output_dir)

        # Add a cover sheet
        self.add_cover_sheet(wb, processed_df, results_data)
        
        # Add _excluded sheet at the end if excluded_df is provided and not empty
        if excluded_df is not None and not excluded_df.empty:
            ws_excluded = wb.create_sheet(title="_excluded")
            for col_idx, col in enumerate(excluded_df.columns, 1):
                ws_excluded.cell(row=1, column=col_idx, value=col)
            for row_idx, row in enumerate(excluded_df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    ws_excluded.cell(row=row_idx, column=col_idx, value=value)
        
        # Save workbook
        timestamp = format_datetime(datetime.datetime.now())
        output_path = os.path.join(
            self.output_dir, f"Enhanced_Vulnerability_Analysis_{timestamp}.xlsx"
        )
        
        try:
            wb.save(output_path)
            logger.info(f"Enhanced report saved to {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Error saving enhanced report: {str(e)}")
            return None
    
    def add_cover_sheet(self, wb, processed_df, results_data):
        """
        Add a professional cover sheet to the workbook.
        
        Args:
            wb: Workbook
            processed_df: Processed vulnerability DataFrame
            results_data: Dictionary with analysis results
        """
        # Create cover sheet
        ws = wb.create_sheet(title="Cover", index=0)
        
        # Set column widths
        for col in range(1, 20):
            col_letter = chr(64 + col) if col <= 26 else chr(64 + col // 26) + chr(64 + col % 26)
            ws.column_dimensions[col_letter].width = 12
        
        # Adjust specific columns
        ws.column_dimensions['A'].width = 5  # Margin
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
        
        # Set row heights
        ws.row_dimensions[1].height = 50  # Top margin
        
        # Add version number
        version_row = 2
        version_cell = ws.cell(row=version_row, column=2, value="Tool Version: 2.1.4")
        version_cell.font = Font(size=12, italic=True, color=self.kpmg_dark_blue)
        ws.merge_cells(start_row=version_row, start_column=2, end_row=version_row, end_column=8)
        ws.row_dimensions[version_row].height = 20

        # Title section with KPMG styling
        title_row = 3
        title_cell = ws.cell(row=title_row, column=2, value="VULNERABILITY ASSESSMENT REPORT")
        title_cell.font = Font(size=24, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color=self.kpmg_blue, end_color=self.kpmg_blue, fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=title_row, start_column=2, end_row=title_row, end_column=12)
        ws.row_dimensions[title_row].height = 40
        
        # Subtitle
        subtitle_row = title_row + 1
        subtitle_cell = ws.cell(row=subtitle_row, column=2, value="Security Analysis & Remediation Insights")
        subtitle_cell.font = Font(size=16, italic=True, color="FFFFFF")
        subtitle_cell.fill = PatternFill(start_color=self.kpmg_dark_blue, end_color=self.kpmg_dark_blue, fill_type="solid")
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=subtitle_row, start_column=2, end_row=subtitle_row, end_column=12)
        ws.row_dimensions[subtitle_row].height = 30
        
        # Date section
        date_row = subtitle_row + 2
        date_label = ws.cell(row=date_row, column=2, value="Report Date:")
        date_label.font = Font(bold=True)
        
        date_value = ws.cell(row=date_row, column=3, value=datetime.datetime.now().strftime("%d %B %Y"))
        date_value.font = Font(bold=True)
        
        # Executive summary section
        summary_title_row = date_row + 2
        summary_title = ws.cell(row=summary_title_row, column=2, value="EXECUTIVE SUMMARY")
        summary_title.font = Font(size=14, bold=True, color=self.kpmg_dark_blue)
        
        # Calculate summary statistics
        total_vulns = len(processed_df) if processed_df is not None else 0
        risk_counts = {}
        
        if processed_df is not None and "Risk" in processed_df.columns:
            risk_counts = processed_df["Risk"].value_counts().to_dict()
        
        critical_count = risk_counts.get("Critical", 0)
        high_count = risk_counts.get("High", 0)
        medium_count = risk_counts.get("Medium", 0)
        low_count = risk_counts.get("Low", 0)
        
        # Format statistics
        host_count = len(processed_df["Host"].unique()) if processed_df is not None and "Host" in processed_df.columns else 0
        kev_count = len(processed_df[processed_df["KEV Listed"] == "Yes"]) if processed_df is not None and "KEV Listed" in processed_df.columns else 0
        
        # Write executive summary
        summary_row = summary_title_row + 1
        summary_text = (
            f"This report presents a comprehensive vulnerability assessment of {host_count} hosts, "
            f"identifying a total of {total_vulns} vulnerabilities. "
            f"The analysis reveals {critical_count} Critical, {high_count} High, {medium_count} Medium, and {low_count} Low risk vulnerabilities. "
            f"Of particular concern are {kev_count} vulnerabilities that appear on the CISA Known Exploited Vulnerabilities (KEV) catalog, "
            f"indicating active exploitation in the wild. "
            f"\n\nA risk-based prioritization approach is recommended, focusing first on Critical vulnerabilities and those "
            f"on the KEV list, followed by vulnerabilities with high EPSS scores indicating likelihood of exploitation. "
            f"This report provides detailed insights, visualizations, and actionable remediation guidance for all identified issues."
        )
        
        summary_cell = ws.cell(row=summary_row, column=2, value=summary_text)
        summary_cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=summary_row, start_column=2, end_row=summary_row+6, end_column=10)
        
        # Add KPMG-styled border around the executive summary
        border = Border(
            left=Side(style='medium', color=self.kpmg_dark_blue),
            right=Side(style='medium', color=self.kpmg_dark_blue),
            top=Side(style='medium', color=self.kpmg_dark_blue),
            bottom=Side(style='medium', color=self.kpmg_dark_blue)
        )
        
        for row in range(summary_row, summary_row+7):
            for col in range(2, 11):
                ws.cell(row=row, column=col).border = border
        
        # Add key statistics section
        stats_row = summary_row + 8
        stats_title = ws.cell(row=stats_row, column=2, value="KEY STATISTICS")
        stats_title.font = Font(size=14, bold=True, color=self.kpmg_dark_blue)
        
        # Create a 2x4 grid of key statistics with KPMG styling
        stats_data = [
            ["Total Vulnerabilities", str(total_vulns)],
            ["Critical Vulnerabilities", str(critical_count)],
            ["High Vulnerabilities", str(high_count)],
            ["Medium Vulnerabilities", str(medium_count)],
            ["Low Vulnerabilities", str(low_count)],
            ["Hosts Analyzed", str(host_count)],
            ["KEV Listed Vulnerabilities", str(kev_count)],
            ["Report Pages", str(len(wb.sheetnames))]
        ]
        
        # Write statistics in a 2-column grid
        for i, (label, value) in enumerate(stats_data):
            row = stats_row + 1 + (i // 2)
            col_start = 2 + (i % 2) * 5
            
            # Label
            label_cell = ws.cell(row=row, column=col_start, value=label)
            label_cell.font = Font(bold=True)
            label_cell.fill = PatternFill(start_color=self.kpmg_light_blue, end_color=self.kpmg_light_blue, fill_type="solid")
            label_cell.alignment = Alignment(horizontal="right")
            label_cell.border = border
            
            # Value
            value_cell = ws.cell(row=row, column=col_start+1, value=value)
            value_cell.font = Font(bold=True)
            
            # Color code values based on content
            if "Critical" in label and int(value) > 0:
                value_cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                value_cell.font = Font(bold=True, color="FFFFFF")
            elif "High" in label and int(value) > 0:
                value_cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
            elif "Medium" in label and int(value) > 0:
                value_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            elif "Low" in label and int(value) > 0:
                value_cell.fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
                value_cell.font = Font(bold=True, color="FFFFFF")
            elif "KEV" in label and int(value) > 0:
                value_cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                value_cell.font = Font(bold=True, color="FFFFFF")
            else:
                value_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            value_cell.border = border
            
            # Merge cells for better layout
            ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_start+2)
            ws.merge_cells(start_row=row, start_column=col_start+1, end_row=row, end_column=col_start+3)
        
        # Report navigation section
        nav_row = stats_row + 6
        nav_title = ws.cell(row=nav_row, column=2, value="REPORT NAVIGATION")
        nav_title.font = Font(size=14, bold=True, color=self.kpmg_dark_blue)
        
        # Add sheet descriptions
        sheet_desc = [
            ["Summary", "Dashboard overview of vulnerability statistics and trends"],
            ["1. Prioritization Insights", "Consolidated view of most critical issues for remediation"],
            ["1.1 Exploitability Scoring", "Detailed scoring of vulnerabilities by exploitation potential"],
            ["1.2 Ageing of Vulnerability", "Analysis of vulnerability age since discovery"],
            ["1.3 Most Exploitable", "Hosts with highest exploitability risk score"],
        ]
        
        # Write sheet navigation
        for i, (sheet_name, desc) in enumerate(sheet_desc):
            row = nav_row + 1 + i
            
            # Sheet name
            name_cell = ws.cell(row=row, column=2, value=sheet_name)
            name_cell.font = Font(bold=True)
            name_cell.fill = PatternFill(start_color=self.kpmg_light_blue, end_color=self.kpmg_light_blue, fill_type="solid")
            name_cell.border = border
            
            # Description
            desc_cell = ws.cell(row=row, column=3, value=desc)
            desc_cell.border = border
            
            # Merge description cells
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
        
        # Footer section
        footer_row = nav_row + len(sheet_desc) + 3
        footer_cell = ws.cell(row=footer_row, column=2, value="© KPMG Vulnerability Assessment Tool")
        footer_cell.font = Font(italic=True, color=self.kpmg_dark_blue)
        ws.merge_cells(start_row=footer_row, start_column=2, end_row=footer_row, end_column=10)

        # Add key enhancements section
        enhancements_row = footer_row + 2
        enhancements_title = ws.cell(row=enhancements_row, column=2, value="KEY ENHANCEMENTS")
        enhancements_title.font = Font(size=14, bold=True, color=self.kpmg_dark_blue)
        ws.merge_cells(start_row=enhancements_row, start_column=2, end_row=enhancements_row, end_column=10)
        enhancements_list = [
            "• Improved risk color palette for better visibility",
            "• Enhanced cover page with version and summary",
            "• Dashboard and summary layout improvements",
            "• Separate workbook for 'Check Needed' vulnerabilities",
            "• Improved sheet order and navigation",
            "• Section dividers and visual consistency across sheets"
        ]
        for i, enh in enumerate(enhancements_list, enhancements_row+1):
            cell = ws.cell(row=i, column=2, value=enh)
            cell.font = Font(size=12)
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=10)
    
    def create_check_needed_workbook(self, check_needed_df):
        """
        Create a separate workbook for 'Check Needed' vulnerabilities (name, host, plugin ID).
        """
        if check_needed_df is None or check_needed_df.empty:
            return None
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Manual Revision Needed"
        # Write headers
        headers = ["Name", "Host", "Plugin ID"]
        ws.append(headers)
        # Write data
        for _, row in check_needed_df.iterrows():
            ws.append([
                row.get("Name", ""),
                row.get("Host", ""),
                row.get("Plugin ID", "")
            ])
        # Save workbook
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"manual_revision_{timestamp}.xlsx"
        output_path = os.path.join(self.output_dir, filename)
        wb.save(output_path)
        logger.info(f"Manual revision workbook saved to {output_path}")
        return output_path

    def generate_all_reports(self, original_df, processed_df, check_needed_df, analyzed_df, results_data, excluded_df=None):
        """
        Generate all report types.
        
        Args:
            original_df: Original vulnerability DataFrame
            processed_df: Processed vulnerability DataFrame
            check_needed_df: Vulnerabilities that need checking
            analyzed_df: Analyzed vulnerability DataFrame
            results_data: Dictionary with analysis results
        
        Returns:
            Tuple of (standard_report_path, json_report_path, enhanced_report_path, check_needed_workbook_path)
        """
        logger.info("Generating all reports")
        
        # Generate standard Excel report
        standard_report = self.create_standard_report(
            original_df, processed_df, check_needed_df, analyzed_df, results_data
        )
        
        # Generate JSON report
        json_report = self.create_json_report(results_data)
        
        # Generate enhanced Excel report
        enhanced_report = self.create_enhanced_report(
            original_df, processed_df, analyzed_df, results_data, excluded_df=excluded_df
        )
        
        # Generate manual revision workbook for 'Check Needed' items
        check_needed_workbook = self.create_check_needed_workbook(check_needed_df)
        
        return standard_report, json_report, enhanced_report, check_needed_workbook