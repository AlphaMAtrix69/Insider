"""Vulnerable Hosts Summary sheet generator."""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from va_tool.reporting.sheets.base import BaseSheetGenerator


class HostsSummarySheetGenerator(BaseSheetGenerator):
    """Generator for the 5.2 Vulnerable Hosts Summary sheet."""

    def __init__(self):
        super().__init__(title="5.2 Vulnerable Hosts Summary")

    def generate(self, wb, df=None, output_dir=None, **kwargs):
        self.logger.info("Generating Vulnerable Hosts Summary sheet")
        ws = super().generate(wb)

        # Add title
        self.add_title(ws, "Hosts Summary", font_size=14, merge_range='A1:G1')

        if df is not None and 'Host' in df.columns and 'Risk' in df.columns:
            df = df[df['Risk'] != 'None']

            # Vulnerability count per host
            vuln_counts = df['Host'].value_counts().reset_index()
            vuln_counts.columns = ['Host', 'Vulnerability Count']

            # CVE counts per host
            if 'CVE' in df.columns:
                cve_counts = df[df['CVE'].apply(lambda x: pd.notna(x) and x != '')]['Host'].value_counts().reset_index()
                cve_counts.columns = ['Host', 'Vulnerabilities with CVE']
                vuln_summary = vuln_counts.merge(cve_counts, on='Host', how='left').fillna({'Vulnerabilities with CVE': 0})
                vuln_summary['Vulnerabilities with CVE'] = vuln_summary['Vulnerabilities with CVE'].astype(int)
            else:
                vuln_summary = vuln_counts.copy()
                vuln_summary['Vulnerabilities with CVE'] = 0

            # Severity breakdown
            severity_counts = df.copy()
            severity_cols = ['Critical', 'High', 'Medium', 'Low']
            severity_pivot = severity_counts.pivot_table(index='Host', columns='Risk', aggfunc='size', fill_value=0).reset_index()

            for col in severity_cols:
                if col not in severity_pivot.columns:
                    severity_pivot[col] = 0
            severity_pivot = severity_pivot[['Host'] + severity_cols]

            vuln_summary = vuln_summary.merge(severity_pivot, on='Host', how='left').fillna(0)
            for col in severity_cols:
                vuln_summary[col] = vuln_summary[col].astype(int)

            # Sort: highest vuln count then critical count
            vuln_summary = vuln_summary.sort_values(['Vulnerability Count', 'Critical'], ascending=[False, False])

            # Column widths
            ws.column_dimensions['A'].width = 30  # Host
            for col in 'BCDEFG':
                ws.column_dimensions[col].width = 20

            # Headers
            headers = ['Host', 'Vulnerability Count', 'Vulnerabilities with CVE'] + severity_cols
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

            # Data cells
            for row_idx, row in enumerate(vuln_summary.itertuples(index=False), 4):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        else:
            ws['A3'] = "No vulnerability data or 'Host'/'Risk' columns available."
            ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions['A'].width = 40

        return ws
