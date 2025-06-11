"""Unique Vulnerabilities sheet generator."""

import pandas as pd
import matplotlib.pyplot as plt
import os
import numpy as np
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from va_tool.reporting.sheets.base import BaseSheetGenerator
from va_tool.utils import ensure_dir_exists, write_df_to_sheet, style_header_cell


class UniqueVulnSheetGenerator(BaseSheetGenerator):
    """Generator for the Unique Vulnerabilities sheet."""
    
    def __init__(self):
        """Initialize the generator."""
        super().__init__(title="3.3 Unique Vulnerabilities")
        self.risk_colors = {
            "Critical": "FF7D7D",  # Darker pastel red
            "High": "FFB366",      # Darker pastel orange
            "Medium": "FFDA66",    # Darker pastel yellow
            "Low": "99CC99"        # Darker pastel green
        }
    
    def generate(self, wb, df=None, output_dir=None, **kwargs):
        """
        Generate the Unique Vulnerabilities sheet.

        Args:
            wb: Excel workbook
            df: DataFrame with vulnerability data
            output_dir: Directory to save chart images
            **kwargs: Additional arguments

        Returns:
            The worksheet
        """
        self.logger.info("Generating Unique Vulnerabilities sheet")
        ws = super().generate(wb)

        # Add title
        self.add_title(
            ws, "Unique Vulnerabilities Summary", 
            font_size=14, merge_range='A1:N1'
        )

        # Add introduction
        intro = ws.cell(row=3, column=1, value="This sheet provides a high-level summary of unique vulnerabilities by severity, aiding in operational planning and resource allocation.")
        intro.font = Font(italic=True)
        ws.merge_cells('A3:N3')

        # Set column widths
        ws.column_dimensions['A'].width = 35  # Vulnerability Name
        ws.column_dimensions['B'].width = 15  # Severity
        ws.column_dimensions['C'].width = 20  # CVE
        ws.column_dimensions['D'].width = 15  # Affected Hosts
        ws.column_dimensions['E'].width = 15  # Instance Count
        ws.column_dimensions['F'].width = 15  # KEV Listed
        ws.column_dimensions['G'].width = 18  # CVSS Score
        ws.column_dimensions['H'].width = 18  # EPSS Score
        ws.column_dimensions['I'].width = 80  # Remediation

        if df is not None and 'Name' in df.columns:
            # Process unique vulnerabilities
            unique_vulns = self.process_unique_vulnerabilities(df)

            # Add overview section
            self.add_overview_section(ws, df, unique_vulns)

            # Find the last row of merged cells (overview section)
            last_merged_row = 1
            for merged in ws.merged_cells.ranges:
                last_merged_row = max(last_merged_row, merged.max_row)

            # Add a blank row after the overview
            table_start_row = last_merged_row + 2

            # Prepare DataFrame for main table
            table_df = pd.DataFrame(unique_vulns)
            headers = [
                "Vulnerability Name", "Severity", "CVE", 
                "Affected Hosts", "Instance Count", "KEV Listed",
                "CVSS Score", "EPSS Score", "Remediation"
            ]

            # Write table using utility (handles merged cells safely)
            write_df_to_sheet(ws, table_df[headers], start_row=table_start_row, start_col=1, include_header=True, style_header=True)

            # Apply risk color formatting to Severity column
            for row in ws.iter_rows(min_row=table_start_row+1, max_row=table_start_row+len(table_df), min_col=2, max_col=2):
                for cell in row:
                    val = cell.value
                    if val in self.risk_colors:
                        cell.fill = PatternFill(
                            start_color=self.risk_colors[val],
                            end_color=self.risk_colors[val],
                            fill_type="solid"
                        )

            # Apply critical color to KEV Listed 'Yes'
            for row in ws.iter_rows(min_row=table_start_row+1, max_row=table_start_row+len(table_df), min_col=6, max_col=6):
                for cell in row:
                    if cell.value == "Yes":
                        cell.fill = PatternFill(
                            start_color=self.risk_colors["Critical"],
                            end_color=self.risk_colors["Critical"],
                            fill_type="solid"
                        )

            # Borders for all data cells
            for row in ws.iter_rows(min_row=table_start_row, max_row=table_start_row+len(table_df), min_col=1, max_col=9):
                for cell in row:
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
        else:
            ws['A5'] = "No vulnerability data available or 'Name' column missing."

        return ws
    
    def process_unique_vulnerabilities(self, df):
        """Process and extract unique vulnerabilities from the dataset."""
        unique_vulns = []
        
        # Group by vulnerability name
        vulnerability_groups = df.groupby('Name')
        
        for name, group in vulnerability_groups:
            # Determine severity (use most common or highest if tied)
            severity_counts = group['Risk'].value_counts()
            if severity_counts.empty:
                severity = "Unknown"
            else:
                # If there's a tie, use the highest severity
                severity_order = {'Critical': 4, 'High': 3, 'Medium': 2, 'Low': 1, 'Informational': 0}
                if len(severity_counts) > 1 and severity_counts.iloc[0] == severity_counts.iloc[1]:
                    # There's a tie, use highest severity
                    severities = severity_counts.index.tolist()
                    severity = max(severities, key=lambda x: severity_order.get(x, 0))
                else:
                    # Use most common
                    severity = severity_counts.index[0]
            
            # Get CVEs (might be multiple)
            cves = group['CVE'].dropna().unique()
            cve_str = ", ".join(str(cve) for cve in cves[:3])
            if len(cves) > 3:
                cve_str += f" (+{len(cves) - 3} more)"
            
            # Count affected hosts
            affected_hosts = len(group['Host'].unique())
            
            # Count instances
            instance_count = len(group)
            
            # Check if KEV listed
            kev_listed = "Yes" if any(group.get('KEV Listed', '') == 'Yes') else "No"
            
            # Get CVSS score (use highest)
            cvss_score = "N/A"
            if 'CVSS v3.0 Base Score' in group.columns:
                scores = group['CVSS v3.0 Base Score'].dropna()
                if not scores.empty:
                    cvss_score = max(scores)
            
            # Get EPSS score (use highest)
            epss_score = "N/A"
            if 'EPSS Score' in group.columns:
                scores = group['EPSS Score'].dropna()
                if not scores.empty:
                    epss_score = max(scores)
            
            # Get remediation guidance
            remediation = "N/A"
            if 'Solution' in group.columns:
                solutions = group['Solution'].dropna().unique()
                if solutions.any():
                    # Use the most comprehensive solution (usually the longest)
                    remediation = max(solutions, key=len)
                    
                    # Truncate if extremely long
                    if len(remediation) > 500:
                        remediation = remediation[:500] + "..."
            
            unique_vulns.append({
                "Vulnerability Name": name,
                "Severity": severity,
                "CVE": cve_str,
                "Affected Hosts": affected_hosts,
                "Instance Count": instance_count,
                "KEV Listed": kev_listed,
                "CVSS Score": cvss_score,
                "EPSS Score": epss_score,
                "Remediation": remediation
            })
        
        # Sort by severity and then by instance count
        severity_order = {'Critical': 4, 'High': 3, 'Medium': 2, 'Low': 1, 'Informational': 0, 'Unknown': -1}
        unique_vulns.sort(key=lambda x: (severity_order.get(x["Severity"], -1), x["Instance Count"]), reverse=True)
        
        return unique_vulns
    
    def add_overview_section(self, ws, df, unique_vulns):
        """Add an overview section with key metrics."""
        # Calculate metrics
        total_vulns = len(df)
        unique_count = len(unique_vulns)
        
        # Count by severity
        severity_counts = {severity: 0 for severity in ['Critical', 'High', 'Medium', 'Low', 'Informational']}
        for vuln in unique_vulns:
            if vuln["Severity"] in severity_counts:
                severity_counts[vuln["Severity"]] += 1
        
        # Count KEV listed
        kev_count = sum(1 for vuln in unique_vulns if vuln["KEV Listed"] == "Yes")
        
        # Add section title
        overview_title = ws.cell(row=5, column=1, value="Vulnerability Overview")
        overview_title.font = Font(size=12, bold=True, color=self.kpmg_blue)
        ws.merge_cells('A5:D5')
        
        # Create metrics table
        headers = ["Metric", "Value", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.kpmg_dark_blue, end_color=self.kpmg_dark_blue, fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Add metrics rows
        metrics = [
            ["Total Vulnerabilities", total_vulns, ""],
            ["Unique Vulnerabilities", unique_count, "100%"],
            ["Critical Vulnerabilities", severity_counts["Critical"], f"{(severity_counts['Critical']/unique_count*100):.1f}%" if unique_count > 0 else "0%"],
            ["High Vulnerabilities", severity_counts["High"], f"{(severity_counts['High']/unique_count*100):.1f}%" if unique_count > 0 else "0%"],
            ["Medium Vulnerabilities", severity_counts["Medium"], f"{(severity_counts['Medium']/unique_count*100):.1f}%" if unique_count > 0 else "0%"],
            ["Low Vulnerabilities", severity_counts["Low"], f"{(severity_counts['Low']/unique_count*100):.1f}%" if unique_count > 0 else "0%"],
            ["KEV Listed Vulnerabilities", kev_count, f"{(kev_count/unique_count*100):.1f}%" if unique_count > 0 else "0%"]
        ]
        
        for row_idx, metric in enumerate(metrics, 7):
            for col_idx, value in enumerate(metric, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Apply conditional formatting based on metric
                if col_idx == 1:  # Metric name
                    cell.font = Font(bold=True)
                    
                    # Add background color based on the metric
                    if "Critical" in metric[0]:
                        cell.fill = PatternFill(start_color=self.risk_colors["Critical"], end_color=self.risk_colors["Critical"], fill_type="solid")
                    elif "High" in metric[0]:
                        cell.fill = PatternFill(start_color=self.risk_colors["High"], end_color=self.risk_colors["High"], fill_type="solid")
                    elif "Medium" in metric[0]:
                        cell.fill = PatternFill(start_color=self.risk_colors["Medium"], end_color=self.risk_colors["Medium"], fill_type="solid")
                    elif "Low" in metric[0]:
                        cell.fill = PatternFill(start_color=self.risk_colors["Low"], end_color=self.risk_colors["Low"], fill_type="solid")
                    elif "KEV" in metric[0]:
                        cell.fill = PatternFill(start_color=self.risk_colors["Critical"], end_color=self.risk_colors["Critical"], fill_type="solid")
                
                # Add borders to all cells
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Add explanation of the data
        explanation_row = row_idx + 1
        explanation = "Note: This table shows unique vulnerabilities by type, allowing for better understanding of the security posture and required remediation efforts."
        cell = ws.cell(row=explanation_row, column=1, value=explanation)
        cell.font = Font(italic=True)
        ws.merge_cells(start_row=explanation_row, start_column=1, end_row=explanation_row, end_column=4)