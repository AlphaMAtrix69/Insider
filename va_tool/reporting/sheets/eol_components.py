"""EOL Components sheet generator with enhanced Excel visualizations."""

import pandas as pd
import os
import re
from datetime import datetime, timedelta
from openpyxl.chart import BarChart, Reference, PieChart, LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import Series
from openpyxl.chart.marker import Marker
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.drawing.colors import ColorMapping
from va_tool.reporting.sheets.base import BaseSheetGenerator
from va_tool.utils import ensure_dir_exists


class EOLComponentsSheetGenerator(BaseSheetGenerator):
    """Generator for the EOL Components sheet."""
    
    def __init__(self):
        """Initialize the generator."""
        super().__init__(title="6.1 EOL Components")
    
    def generate(self, wb, df=None, output_dir=None, results_data=None, **kwargs):
        """
        Generate the EOL Components sheet.
        
        Args:
            wb: Excel workbook
            df: DataFrame with vulnerability data
            output_dir: Directory to save chart images
            results_data: Dictionary with analysis results
            **kwargs: Additional arguments
        
        Returns:
            The worksheet
        """
        self.logger.info("Generating EOL Components sheet")
        ws = super().generate(wb)
        
        # Add title
        self.add_title(
            ws, "Total Components in SEoL", 
            font_size=14, merge_range='A1:F1'
        )
        
        # Load data from results_data if available
        seol_count = 0
        seol_items = None
        
        # First, try to get the exact SEOL items that were reported in the standard report
        if results_data:
            # If we have full analyzed_df in the provided dataframe and results_data has common_critical,
            # we're likely working with the same dataframe used by the standard report
            if df is not None and 'common_critical' in results_data:
                # Use the exact same filter as the standard report
                seol_items = df[df["Name"].str.contains("SEoL", case=False, na=False)]
                # Get unique count of Plugin ID + Name combinations to avoid counting the same vulnerability
                # on different hosts as separate items
                seol_unique = seol_items.drop_duplicates(subset=['Plugin ID', 'Name'])
                seol_count = len(seol_unique)
                
                self.logger.info(f"Found {seol_count} unique SEoL components from {len(seol_items)} total instances")
            
            # If the above doesn't yield the right count, try the seol_summary if available
            elif 'seol_summary' in results_data and not seol_count:
                seol_summary = results_data.get('seol_summary', [])
                
                if isinstance(seol_summary, list) and seol_summary:
                    # If it's a list of dictionaries with SEoL Count
                    if 'SEoL Count' in seol_summary[0]:
                        total_from_summary = sum(item.get('SEoL Count', 0) for item in seol_summary)
                        self.logger.info(f"seol_summary indicates {total_from_summary} SEoL items")
                
                # If we still need to find SEoL items from df
                if df is not None and seol_count == 0:
                    # Try once more with the exact filter
                    seol_items = df[df["Name"].str.contains("SEoL", case=False, na=False)]
                    seol_unique = seol_items.drop_duplicates(subset=['Plugin ID', 'Name'])
                    seol_count = len(seol_unique)
                    self.logger.info(f"Second attempt: found {seol_count} unique SEoL components")
        
        # If we still don't have a count and have a dataframe, try filtering it
        if seol_count == 0 and df is not None:
            # Use the exact same filter as the standard report
            seol_items = df[df["Name"].str.contains("SEoL", case=False, na=False)]
            seol_unique = seol_items.drop_duplicates(subset=['Plugin ID', 'Name'])
            seol_count = len(seol_unique)
            self.logger.info(f"Fallback: found {seol_count} unique SEoL components")
        
        # Now use the data for the sheet
        if seol_count > 0 and seol_unique is not None:
            # Write total SEoL components (using the unique count)
            ws['A3'] = "Total Unique SEoL Components:"
            ws['B3'] = seol_count
            ws['A3'].font = ws['A3'].font.copy(bold=True)
            
            # Also show total instances if different
            if len(seol_items) > seol_count:
                ws['A4'] = "Total SEoL Instances (across all hosts):"
                ws['B4'] = len(seol_items)
            
            # Highlight the total with a fill color
            ws['B3'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws['B3'].font = Font(bold=True, size=12, color="9C0006")
            
            # Process data for duration summary
            eol_with_duration = self.calculate_eol_duration(seol_unique)
            
            # Summarize by duration buckets
            duration_summary = self.summarize_by_duration(eol_with_duration)
            
            # Store chart positions to avoid overlaps
            current_row = 6  # Starting row for duration summary
            if 'B4' in ws:
                current_row = 7  # Adjust if we have the "Total Instances" row
            
            # Write duration summary table and get next row position
            next_row = self.write_duration_summary(ws, duration_summary.to_dict('records'), current_row)
            
            # Create Excel charts for the duration summary with dynamic positioning
            chart_positions = self.create_enhanced_visualizations_from_df(
                ws, duration_summary, eol_with_duration, 
                start_row=current_row, 
                end_row=next_row
            )
            
            # Write detailed SEoL components list after all charts
            components_start_row = chart_positions['end_row'] + 2
            self.write_eol_details_from_df(ws, eol_with_duration, start_row=components_start_row)
            
        else:
            ws['A3'] = "No SEoL components found in the vulnerability data."
        
        return ws
    
    def calculate_eol_duration(self, eol_df):
        """Calculate how long components have been EOL."""
        # Create a copy to avoid modifying the original
        result_df = eol_df.copy()
        
        # Current date for calculation
        current_date = datetime.now().date()
        
        # If CVE Published Date is available, use it to calculate duration
        if 'CVE Published Date' in result_df.columns:
            result_df['EOL Duration Days'] = result_df['CVE Published Date'].apply(
                lambda x: (current_date - datetime.fromisoformat(x).date()).days 
                if pd.notna(x) and x else None
            )
        else:
            # Otherwise, use Days After Discovery if available
            if 'Days After Discovery' in result_df.columns:
                result_df['EOL Duration Days'] = result_df['Days After Discovery']
            else:
                result_df['EOL Duration Days'] = None
        
        return result_df
    
    def summarize_by_duration(self, eol_df):
        """Summarize EOL components by duration buckets."""
        # Create duration buckets
        def categorize_duration(days):
            if pd.isna(days):
                return "Unknown"
            elif days <= 30:
                return "< 30 days"
            elif days <= 90:
                return "30-90 days"
            elif days <= 180:
                return "90-180 days"
            elif days <= 365:
                return "180-365 days"
            else:
                return "> 365 days"
        
        eol_df['Duration Category'] = eol_df['EOL Duration Days'].apply(categorize_duration)
        
        # Count components by duration category
        duration_summary = eol_df['Duration Category'].value_counts().reset_index()
        duration_summary.columns = ['Duration', 'Component Count']
        
        # Set priority order for categories
        duration_order = ["< 30 days", "30-90 days", "90-180 days", "180-365 days", "> 365 days", "Unknown"]
        duration_summary['Order'] = duration_summary['Duration'].apply(lambda x: duration_order.index(x))
        duration_summary = duration_summary.sort_values('Order').drop('Order', axis=1)
        
        return duration_summary
    
    def write_duration_summary(self, ws, duration_summary, start_row):
        """
        Write duration summary table to worksheet.
        
        Args:
            ws: Worksheet
            duration_summary: List of dictionaries with duration summary data
            start_row: Starting row for the table
            
        Returns:
            The next available row after the table
        """
        # Write headers with styling
        ws.cell(row=start_row, column=1, value="EOL Duration") 
        ws.cell(row=start_row, column=2, value="Component Count")
        
        # Apply header styling
        for col in [1, 2]:
            cell = ws.cell(row=start_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data with alternate row coloring
        row_idx = start_row + 1
        total_count = 0
        
        for i, item in enumerate(duration_summary):
            # Alternate row colors
            fill_color = "E9F1FB" if i % 2 == 0 else "FFFFFF"
            
            # Write values
            ws.cell(row=row_idx, column=1, value=item.get('Duration', ''))
            count = item.get('Component Count', 0)
            ws.cell(row=row_idx, column=2, value=count)
            
            # Apply cell styling
            for col in [1, 2]:
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                if col == 2:  # Right-align count values
                    cell.alignment = Alignment(horizontal="right")
            
            total_count += count
            row_idx += 1
        
        # Calculate total with special styling
        total_row = row_idx
        ws.cell(row=total_row, column=1, value="Total")
        ws.cell(row=total_row, column=2, value=total_count)
        
        # Style the total row
        for col in [1, 2]:
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            if col == 2:  # Right-align count values
                cell.alignment = Alignment(horizontal="right")
        
        # Return the next available row
        return total_row + 1
    
    def write_eol_details_from_df(self, ws, eol_df, start_row):
        """
        Write detailed EOL components list from DataFrame to worksheet.
        
        Args:
            ws: Worksheet
            eol_df: DataFrame with EOL components data
            start_row: Starting row for the detailed list
            
        Returns:
            The next available row after the detailed list
        """
        # Add section title
        self.add_section_title(ws, "Unique SEoL Components", cell=f"A{start_row}")
        
        # Write headers with styling
        headers = ['Plugin ID', 'Name', 'Risk', 'EOL Duration (Days)', 'CVE']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row+1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Sort by duration descending
        if 'EOL Duration Days' in eol_df.columns:
            eol_df = eol_df.sort_values('EOL Duration Days', ascending=False, na_position='last')
        
        # Write data with conditional formatting for risk level
        for i, (_, row) in enumerate(eol_df.iterrows(), 0):
            row_idx = start_row + 2 + i
            
            # Alternate row colors
            fill_color = "E9F1FB" if i % 2 == 0 else "FFFFFF"
            
            # Write and style each cell
            ws.cell(row=row_idx, column=1, value=str(row.get('Plugin ID', ''))).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            ws.cell(row=row_idx, column=2, value=str(row.get('Name', ''))).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            # Apply risk-based formatting
            risk_cell = ws.cell(row=row_idx, column=3, value=str(row.get('Risk', '')))
            risk_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            risk = row.get('Risk', '')
            if risk == 'Critical':
                risk_cell.font = Font(color="FF0000", bold=True)  # Red
            elif risk == 'High':
                risk_cell.font = Font(color="FF8000")  # Orange
            elif risk == 'Medium':
                risk_cell.font = Font(color="FFBF00")  # Amber
            
            # Add duration and CVE
            ws.cell(row=row_idx, column=4, value=row.get('EOL Duration Days', '')).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            ws.cell(row=row_idx, column=5, value=str(row.get('CVE', ''))).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Set column widths
        ws.column_dimensions['A'].width = 15  # Plugin ID
        ws.column_dimensions['B'].width = 50  # Name
        ws.column_dimensions['C'].width = 15  # Risk
        ws.column_dimensions['D'].width = 20  # EOL Duration
        ws.column_dimensions['E'].width = 20  # CVE
        
        # Return the next available row
        return start_row + 2 + len(eol_df)
    
    def create_enhanced_visualizations_from_df(self, ws, duration_summary, eol_df, start_row, end_row):
        """
        Create multiple Excel charts for enhanced visualization from DataFrames with dynamic positioning.
        
        Args:
            ws: Worksheet
            duration_summary: DataFrame with duration summary data
            eol_df: DataFrame with EOL components data
            start_row: Starting row for duration summary table
            end_row: Ending row for duration summary table
            
        Returns:
            Dictionary with chart positions
        """
        # Calculate chart positions
        title_row = start_row - 1  # Row for chart titles
        chart_data_height = end_row - start_row  # Height of the data table
        
        # Bar chart position
        bar_chart_cell = "D5"  # Default position
        # Adjust if needed based on data size
        if chart_data_height > 8:
            bar_chart_cell = f"D{title_row}"
        
        # Create bar chart for duration summary
        self.create_duration_bar_chart_from_df(ws, duration_summary, title_row=title_row, position=bar_chart_cell)
        
        # Pie chart will be placed to the right of the bar chart
        pie_chart_cell = f"I{title_row}"
        
        # Create pie chart for duration distribution
        pie_chart_row = self.create_duration_pie_chart_from_df(ws, duration_summary, position=pie_chart_cell)
        
        # The risk distribution chart will be placed below the bar chart or pie chart
        # Calculate the position based on the larger of the two charts
        bar_chart_end_row = title_row + 12  # Approximate end row for bar chart
        risk_chart_row = max(bar_chart_end_row, pie_chart_row) + 2
        
        # Create risk distribution chart if risk data is available
        risk_end_row = risk_chart_row
        if 'Risk' in eol_df.columns:
            risk_end_row = self.create_risk_distribution_chart_from_df(
                ws, eol_df, 
                title_row=risk_chart_row, 
                position=f"D{risk_chart_row}"
            )
        
        # Return the positions for further layout calculations
        return {
            'bar_chart': {'start': title_row, 'end': bar_chart_end_row},
            'pie_chart': {'start': title_row, 'end': pie_chart_row},
            'risk_chart': {'start': risk_chart_row, 'end': risk_end_row},
            'end_row': risk_end_row + 2  # Add some padding
        }
    
    def create_duration_bar_chart_from_df(self, ws, duration_df, title_row, position):
        """
        Create an enhanced bar chart for duration summary from DataFrame.
        
        Args:
            ws: Worksheet
            duration_df: DataFrame with duration summary data
            title_row: Row for chart title
            position: Cell position for chart
            
        Returns:
            End row for the chart
        """
        # Write chart title
        ws.cell(row=title_row, column=4, value="SEoL Components by Duration")
        ws.cell(row=title_row, column=4).font = Font(bold=True, size=12)
        
        # Create bar chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 42  # Use a more colorful Excel chart style
        chart.title = None  # We already added a title in the worksheet
        chart.y_axis.title = "Number of Components"
        chart.x_axis.title = "Duration Since SEoL"
        
        # Find the data range
        header_row = title_row + 1  # For the header of the duration table
        data_row_count = len(duration_df) + 1  # +1 for header
        
        # Create references to the data
        data = Reference(ws, min_col=2, max_col=2, min_row=header_row, max_row=header_row+data_row_count)
        cats = Reference(ws, min_col=1, max_col=1, min_row=header_row+1, max_row=header_row+data_row_count)
        
        # Add data and categories
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Enable data labels
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showCatName = False
        
        # Set series colors
        s = chart.series[0]
        s.graphicalProperties.solidFill = "5B9BD5"  # Blue fill
        
        # Set chart size and position - adjust height based on data
        chart.width = 15
        chart.height = 8
        
        # Add chart to worksheet
        ws.add_chart(chart, position)
        
        # Return the approximate end row for the chart
        return title_row + chart.height + 2
    
    def create_duration_pie_chart_from_df(self, ws, duration_df, position):
        """
        Create a pie chart showing distribution of components by duration from DataFrame.
        
        Args:
            ws: Worksheet
            duration_df: DataFrame with duration summary data
            position: Cell position for chart
            
        Returns:
            End row for the chart
        """
        # Prepare data for the chart - write to a different area of the sheet
        chart_data_row = 30
        
        # Write headers for source data (hidden area of the sheet)
        ws.cell(row=chart_data_row, column=8, value="Duration")
        ws.cell(row=chart_data_row, column=9, value="Count")
        
        # Write source data for the chart
        for i, (_, row) in enumerate(duration_df.iterrows(), 1):
            ws.cell(row=chart_data_row+i, column=8, value=row['Duration'])
            ws.cell(row=chart_data_row+i, column=9, value=row['Component Count'])
        
        # Create pie chart
        pie = PieChart()
        pie.title = "SEoL Components Distribution"
        pie.style = 10
        
        # Define data range for the chart
        labels = Reference(ws, min_col=8, min_row=chart_data_row+1, max_row=chart_data_row+len(duration_df))
        data = Reference(ws, min_col=9, min_row=chart_data_row, max_row=chart_data_row+len(duration_df))
        
        # Add data to the chart
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        
        # Add data labels showing percentages
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        
        # Set chart size and position
        pie.width = 10
        pie.height = 8
        
        # Add chart to worksheet
        ws.add_chart(pie, position)
        
        # Return the approximate end row for the chart
        chart_row = int(position[1:]) if position[1:].isdigit() else 5
        return chart_row + pie.height + 2
    
    def create_risk_distribution_chart_from_df(self, ws, eol_df, title_row, position):
        """
        Create a chart showing risk distribution of SEoL components from DataFrame.
        
        Args:
            ws: Worksheet
            eol_df: DataFrame with EOL components data
            title_row: Row for chart title
            position: Cell position for chart
            
        Returns:
            End row for the chart
        """
        # Prepare data - count by risk level
        risk_counts = eol_df['Risk'].value_counts().reset_index()
        risk_counts.columns = ['Risk', 'Count']
        
        # Write data to worksheet - use a hidden area
        chart_data_row = 45
        ws.cell(row=chart_data_row, column=8, value="Risk Level")
        ws.cell(row=chart_data_row, column=9, value="Count")
        
        for i, (_, row) in enumerate(risk_counts.iterrows(), 1):
            ws.cell(row=chart_data_row+i, column=8, value=row['Risk'])
            ws.cell(row=chart_data_row+i, column=9, value=row['Count'])
        
        # Create bar chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 42
        chart.title = "SEoL Components by Risk Level"
        chart.y_axis.title = "Number of Components"
        
        # Define data range for the chart
        data = Reference(ws, min_col=9, min_row=chart_data_row, max_row=chart_data_row+len(risk_counts))
        cats = Reference(ws, min_col=8, min_row=chart_data_row+1, max_row=chart_data_row+len(risk_counts))
        
        # Add data to the chart
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Add data labels
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        
        # Color the bars based on risk level
        for i, (risk, _) in enumerate(risk_counts.iterrows()):
            if i < len(chart.series):
                s = chart.series[i]
                risk_name = risk_counts.iloc[i]['Risk']
                if risk_name == "Critical":
                    s.graphicalProperties.solidFill = "FF0000"  # Red
                elif risk_name == "High":
                    s.graphicalProperties.solidFill = "FF8000"  # Orange
                elif risk_name == "Medium":
                    s.graphicalProperties.solidFill = "FFBF00"  # Amber
                else:
                    s.graphicalProperties.solidFill = "00B050"  # Green for Low
        
        # Set chart size and position
        chart.width = 15
        chart.height = 8
        
        # Add chart to worksheet
        ws.add_chart(chart, position)
        
        # Return the approximate end row for the chart
        chart_row = int(position[1:]) if position[1:].isdigit() else title_row
        return chart_row + chart.height + 2