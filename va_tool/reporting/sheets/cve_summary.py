"""CVE Summary sheet generator with full details."""

import pandas as pd
from collections import Counter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from va_tool.reporting.sheets.base import BaseSheetGenerator


class CveSummarySheetGenerator(BaseSheetGenerator):
    """Generator for the 5.1 CVE Summary sheet."""

    def __init__(self):
        super().__init__(title="5.1 CVE Summary")

    def generate(self, wb, df=None, output_dir=None, **kwargs):
        self.logger.info("Generating CVE Summary sheet")
        ws = super().generate(wb)

        # Add title
        self.add_title(ws, "CVE Summary", font_size=14)

        try:
            if df is None or not {'CVE', 'Risk', 'Name', 'Description', 'Host', 'Solution'}.issubset(df.columns):
                ws['A3'] = "Missing required columns in data: CVE, Risk, Name, Description, Host, Solution"
                return ws

            df_cves = df[['CVE', 'Risk', 'Name', 'Description', 'Host', 'Solution']].dropna(subset=['CVE'])
            cve_details = {}
            severity_precedence = {'Critical': 4, 'High': 3, 'Medium': 2, 'Low': 1, 'None': 1, 'Low/None': 1}
            all_cves = []

            for _, row in df_cves.iterrows():
                cves = [cve.strip() for cve in str(row["CVE"]).split(',') if cve.startswith("CVE")]
                severity = 'Low/None' if row["Risk"] in ['Low', 'None'] else row["Risk"]

                for cve in cves:
                    all_cves.append(cve)
                    if cve not in cve_details:
                        cve_details[cve] = {
                            "count": 0,
                            "severity": severity,
                            "name": row["Name"] if pd.notna(row["Name"]) else "N/A",
                            "description": row["Description"] if pd.notna(row["Description"]) else "N/A",
                            "hosts": set(),
                            "solutions": set()
                        }
                    cve_details[cve]["count"] += 1

                    # Update severity if higher
                    if severity_precedence.get(severity, 0) > severity_precedence.get(cve_details[cve]["severity"], 0):
                        cve_details[cve]["severity"] = severity

                    if pd.notna(row["Host"]):
                        cve_details[cve]["hosts"].add(str(row["Host"]))
                    if pd.notna(row["Solution"]):
                        cve_details[cve]["solutions"].add(str(row["Solution"]))

            # Create DataFrame
            summary = []
            for cve, detail in cve_details.items():
                summary.append({
                    "CVE": cve,
                    "Count": detail["count"],
                    "Severity": detail["severity"],
                    "Name": detail["name"],
                    "Description": detail["description"],
                    "Hosts": ", ".join(sorted(detail["hosts"])) if detail["hosts"] else "N/A",
                    "Solutions": ", ".join(sorted(detail["solutions"])) if detail["solutions"] else "N/A"
                })
            summary_df = pd.DataFrame(summary)

            if summary_df.empty:
                ws['A3'] = "No CVEs found in the data."
                return ws

            # Sort by severity > count
            summary_df['SeverityRank'] = summary_df['Severity'].map(severity_precedence)
            summary_df = summary_df.sort_values(by=['SeverityRank', 'Count'], ascending=[False, False])
            summary_df = summary_df.drop(columns=['SeverityRank'])
            summary_df = summary_df.drop_duplicates(subset=['CVE'])

            # Write headers
            headers = ["CVE", "Count", "Severity", "Name", "Description", "Hosts", "Solutions"]
            self.write_headers(ws, headers, row=3)

            for idx, row in enumerate(summary_df.itertuples(index=False), 4):
                ws[f"A{idx}"] = row.CVE
                ws[f"B{idx}"] = row.Count
                ws[f"C{idx}"] = row.Severity
                ws[f"D{idx}"] = row.Name
                ws[f"E{idx}"] = row.Description
                ws[f"F{idx}"] = row.Hosts
                ws[f"G{idx}"] = row.Solutions

            # Set column widths
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 50
            ws.column_dimensions["E"].width = 80
            ws.column_dimensions["F"].width = 50
            ws.column_dimensions["G"].width = 80

        except Exception as e:
            self.logger.error(f"Error generating CVE Summary sheet: {e}")
            ws['A3'] = f"Error generating CVE Summary: {e}"

        return ws
