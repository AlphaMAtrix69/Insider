"""Risk Summary sheet generator with charts and tables."""

import pandas as pd
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from va_tool.reporting.sheets.base import BaseSheetGenerator
from va_tool.utils import ensure_dir_exists
from va_tool.reporting.sheets.top_10_cve_xlsx import Top10CVEsSheetMixin


class RiskSummarySheetGenerator(BaseSheetGenerator, Top10CVEsSheetMixin):
    """Generator for the Risk Summary sheet with charts."""

    def __init__(self):
        super().__init__(title="5. Risk Summary")

    def generate(self, wb, df=None, output_dir=None, **kwargs):
        self.logger.info("Generating Risk Summary sheet")
        ws = super().generate(wb)

        self.add_title(ws, "Vulnerability Risk Summary", font_size=14, merge_range='A1:C1')

        severity_order = ['Critical', 'High', 'Medium', 'Low']

        if df is not None and 'Risk' in df.columns:
            # Normalize Risk values: treat 'None' as 'Low'
            risk_counts = df['Risk'].replace({'None': 'Low'}).value_counts().reset_index()
            risk_counts.columns = ['Severity', 'Count']
            risk_summary = pd.DataFrame({'Severity': severity_order})
            risk_summary = risk_summary.merge(risk_counts, on='Severity', how='left').fillna({'Count': 0})
            risk_summary['Count'] = risk_summary['Count'].astype(int)

            # CVE-related stats
            if 'CVE' in df.columns:
                risk_summary['Vulnerabilities with CVE'] = risk_summary['Severity'].apply(
                    lambda x: df[df['Risk'].isin(['Low', 'None'])]['CVE'].apply(lambda y: pd.notna(y) and y != '').sum()
                    if x == 'Low' else
                    df[df['Risk'] == x]['CVE'].apply(lambda y: pd.notna(y) and y != '').sum()
                )
            else:
                risk_summary['Vulnerabilities with CVE'] = 0

            # Set column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 25

            # Write headers
            headers = ['Severity', 'Count', 'Vulnerabilities with CVE']
            self.write_headers(ws, headers, row=3)

            # Write data
            for row_idx, row in enumerate(risk_summary.itertuples(), 4):
                ws.cell(row=row_idx, column=1, value=row[1])  # Severity
                ws.cell(row=row_idx, column=2, value=row[2])  # Count
                ws.cell(row=row_idx, column=3, value=row[3])  # CVEs

            # Add pie chart
            chart = PieChart()
            chart.title = "Vulnerability Count by Severity"
            chart.title.font = Font(size=16, bold=True)

            # Create adjusted labels for the chart only (Low/None → Low)
            adjusted_labels_col = 1  # Column Z
            adjusted_label_cells = []
            for i in range(4, 8):  # Rows 4–7
                severity = ws.cell(row=i, column=1).value
                adjusted_label = "Low" if severity == "Low/None" else severity
                ws.cell(row=i, column=adjusted_labels_col, value=adjusted_label)
                adjusted_label_cells.append(ws.cell(row=i, column=adjusted_labels_col))

            labels = Reference(ws, min_col=adjusted_labels_col, min_row=4, max_row=7)
            data = Reference(ws, min_col=2, min_row=3, max_row=7)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.height = 10
            chart.width = 15
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.dataLabels.showCatName = False
            chart.dataLabels.showPercent = False
            chart.dataLabels.showSerName = False
            chart.legend.overlay = False

            # Colors
            colors = ['red', 'orange', 'yellow', 'green']
            for i, color in enumerate(colors):
                pt = DataPoint(idx=i)
                pt.graphicalProperties = GraphicalProperties(solidFill=ColorChoice(prstClr=color))
                chart.series[0].data_points.append(pt)

            ws.add_chart(chart, "A10")

        else:
            ws['A3'] = "No vulnerability data or 'Risk' column available."
            ws.column_dimensions['A'].width = 40

        # Top 10 Vulnerable Hosts and chart
        if df is not None and 'Host' in df.columns and 'Risk' in df.columns:
            self.add_title(ws, "Top 10 Vulnerable Hosts", cell='U1', font_size=14, bold=True, merge_range='U1:AA1')

            vuln_counts = df['Host'].value_counts().reset_index()
            vuln_counts.columns = ['Host', 'Vulnerability Count']

            if 'CVE' in df.columns:
                cve_counts = df[df['CVE'].apply(lambda x: pd.notna(x) and x != '')]['Host'].value_counts().reset_index()
                cve_counts.columns = ['Host', 'Vulnerabilities with CVE']
                vuln_summary = vuln_counts.merge(cve_counts, on='Host', how='left').fillna(
                    {'Vulnerabilities with CVE': 0})
                vuln_summary['Vulnerabilities with CVE'] = vuln_summary['Vulnerabilities with CVE'].astype(int)
            else:
                vuln_summary = vuln_counts.copy()
                vuln_summary['Vulnerabilities with CVE'] = 0

            severity_counts = df.copy()
            severity_counts['Risk'] = severity_counts['Risk'].replace('None', 'Low').replace('Low', 'Low')
            severity_cols = ['Critical', 'High', 'Medium', 'Low']
            severity_pivot = severity_counts.pivot_table(index='Host', columns='Risk', aggfunc='size',
                                                         fill_value=0).reset_index()

            for col in severity_cols:
                if col not in severity_pivot.columns:
                    severity_pivot[col] = 0

            severity_pivot = severity_pivot[['Host'] + severity_cols]
            vuln_summary = vuln_summary.merge(severity_pivot, on='Host', how='left').fillna(0)
            for col in severity_cols:
                vuln_summary[col] = vuln_summary[col].astype(int)

            vuln_summary = vuln_summary.sort_values('Vulnerability Count', ascending=False).head(10)

            ws.column_dimensions['U'].width = 30
            for col in 'VWXYZAA':
                ws.column_dimensions[col].width = 20

            headers = ['Host', 'Vulnerability Count', 'Vulnerabilities with CVE'] + severity_cols
            # Manual header writing starting from column U (column 21)
            for col_idx, header in enumerate(headers, 21):
                cell = ws.cell(row=3, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

            for row_idx, row in enumerate(vuln_summary.itertuples(index=False), 4):
                for col_idx, value in enumerate(row, 21):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            chart = BarChart()
            chart.type = "bar"
            chart.grouping = "percentStacked"
            chart.overlap = 100
            chart.title = "Top 10 Hosts by Vulnerability Severity Breakdown"
            chart.x_axis.title = "Host"
            chart.y_axis.title = "Vulnerability Count"
            chart.height = 14
            chart.width = 20
            chart.legend.position = 'b'
            chart.legend.overlay = False
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.dataLabels.showCatName = False
            chart.dataLabels.showPercent = False
            chart.dataLabels.showSerName = False

            data_start_row = 4
            data_end_row = 3 + len(vuln_summary)
            data = Reference(ws, min_col=24, max_col=27, min_row=3, max_row=data_end_row)
            cats = Reference(ws, min_col=21, min_row=4, max_row=data_end_row)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            severity_colors = ['FF0000', 'FFA500', 'FFFF00', '2CA02C']
            for i, series in enumerate(chart.series):
                series.graphicalProperties.solidFill = severity_colors[i]

            ws.add_chart(chart, "U16")

        if df is not None:
            self.generate_top_10_cves(ws, df, output_dir)

        return ws
