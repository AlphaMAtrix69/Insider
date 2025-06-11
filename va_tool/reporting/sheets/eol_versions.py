"""EOL Versions sheet generator that counts and displays all version instances."""

import pandas as pd
import os
import re
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from va_tool.reporting.sheets.base import BaseSheetGenerator
from va_tool.utils import get_logger


class EOLVersionsSheetGenerator(BaseSheetGenerator):
    """Generator for the EOL Versions sheet."""
    
    def __init__(self):
        """Initialize the generator."""
        super().__init__(title="6.3 EOL Versions")
        self.logger = get_logger()
    
    def generate(self, wb, df=None, output_dir=None, results_data=None, **kwargs):
        """Generate the EOL Versions sheet."""
        self.logger.info("Generating EOL Versions sheet")
        ws = super().generate(wb)
        
        # Add title
        self.add_title(
            ws, "Total Different Versions Identified as SEoL", 
            font_size=14, merge_range='A1:F1'
        )
        
        # Extract SEoL data from the main dataframe
        if df is not None:
            # Filter for SEoL entries
            seol_df = df[df["Name"].str.contains("SEoL", case=False, na=False)]
            
            if not seol_df.empty:
                self.logger.info(f"Found {len(seol_df)} SEoL entries")
                
                # Process the SEoL data to extract all version information 
                # without any filtering or truncation
                versions_data = self.process_versions(seol_df)
                
                # Write summary statistics
                self.write_summary_statistics(ws, versions_data)
                
                # Write software distribution with version counts
                self.write_software_distribution(ws, versions_data)
                
                # Create visualizations
                self.create_charts(ws, versions_data)
            else:
                ws['A3'] = "No SEoL data found in the vulnerability data."
        else:
            ws['A3'] = "No vulnerability data provided."
        
        return ws
    
    def process_versions(self, seol_df):
        """
        Process SEoL data to extract all version information.
        Counts total instances of each software and version without any filtering.
        """
        # Create a list to store all version data
        all_versions = []
        
        # Process each SEoL entry
        for _, row in seol_df.iterrows():
            name = str(row['Name']) if pd.notna(row.get('Name')) else ""
            plugin_id = str(row.get('Plugin ID', ''))
            host = str(row.get('Host', ''))
            risk = str(row.get('Risk', ''))
            
            # Extract software type and version from the name
            software_type, version = self.extract_software_and_version(name)
            
            # Add to the versions list
            all_versions.append({
                'Software': software_type,
                'Version': version,
                'Plugin ID': plugin_id,
                'Host': host,
                'Risk': risk,
                'Name': name
            })
        
        # Convert to DataFrame
        versions_df = pd.DataFrame(all_versions)
        
        # Count total instances by Software
        software_counts = versions_df['Software'].value_counts().reset_index()
        software_counts.columns = ['Software', 'Total Instances']
        
        # Count total instances by Software and Version
        version_counts = versions_df.groupby(['Software', 'Version']).size().reset_index(name='Instance Count')
        
        # Count unique versions per software
        unique_versions = versions_df.groupby('Software')['Version'].nunique().reset_index()
        unique_versions.columns = ['Software', 'Unique Versions']
        
        # Merge the counts
        software_summary = pd.merge(software_counts, unique_versions, on='Software')
        
        # Calculate total software and total versions
        total_software = len(software_counts)
        total_versions = len(version_counts)
        total_instances = len(versions_df)
        
        return {
            'versions_df': versions_df,
            'software_counts': software_counts.sort_values('Total Instances', ascending=False),
            'version_counts': version_counts.sort_values(['Software', 'Instance Count'], ascending=[True, False]),
            'software_summary': software_summary.sort_values('Total Instances', ascending=False),
            'total_software': total_software,
            'total_versions': total_versions,
            'total_instances': total_instances
        }
    
    def extract_software_and_version(self, name):
        """Extract software type and version from a SEoL name."""
        # Default values
        software_type = "Unknown"
        version = "Unknown"
        
        # Common software patterns
        software_patterns = [
            # Unsupported Version of X
            (r'Unsupported\s+Version\s+of\s+([\w\.\s\-\(\)]+)', 1),
            # X Unsupported Version
            (r'([\w\.\s\-\(\)]+)\s+Unsupported\s+Version', 1),
            # X SEoL
            (r'([\w\.\s\-\(\)]+)\s+SEoL', 1),
            # SEoL X
            (r'SEoL\s+([\w\.\s\-\(\)]+)', 1),
        ]
        
        # Try each pattern to extract software type
        for pattern, group in software_patterns:
            match = re.search(pattern, name, re.IGNORECASE)
            if match:
                software_type = match.group(group).strip()
                break
        
        # If no match, use the first word after "Unsupported" or "SEoL"
        if software_type == "Unknown":
            clean_name = re.sub(r'^(Unsupported|SEoL)\s+', '', name)
            words = clean_name.split()
            if words:
                software_type = words[0]
        
        # Version patterns
        version_patterns = [
            r'(\d+\.\d+\.\d+(\.\d+)?)',  # Common format x.y.z.w
            r'version\s+(\d+\.\d+(\.\d+)?(\.\d+)?)',  # "version X.Y.Z"
            r'v(\d+\.\d+(\.\d+)?(\.\d+)?)',  # "vX.Y.Z"
            r'(\d+\.\d+)\s',  # "X.Y " (with space after)
            r'(\d+\.\d+[a-z]?)',  # Format x.y or x.ya
            r'Windows\s+Server\s+(\d{4})',  # Windows Server 2012, 2016, etc.
            r'Windows\s+(\d+)',  # Windows 7, 8, 10, 11, etc.
            r'(\d+)\s+SP\d+',  # Version with Service Pack
        ]
        
        # Try each pattern to extract version
        for pattern in version_patterns:
            match = re.search(pattern, name, re.IGNORECASE)
            if match:
                version = match.group(1)
                break
        
        return software_type, version
    
    def write_summary_statistics(self, ws, versions_data):
        """Write summary statistics to the worksheet."""
        ws['A3'] = "SEoL Versions Summary"
        ws['A3'].font = ws['A3'].font.copy(bold=True)
        
        # Display summary statistics with descriptive labels
        ws['A4'] = "Total SEoL Components:"
        ws['B4'] = versions_data['total_instances']
        
        ws['A5'] = "Total Different Software Types:"
        ws['B5'] = versions_data['total_software']
        
        ws['A6'] = "Total Different Versions:"
        ws['B6'] = versions_data['total_versions']
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
    
    def write_software_distribution(self, ws, versions_data):
        """Write software distribution with version counts."""
        # 1. First table: Software summary
        self.write_software_summary_table(ws, versions_data['software_summary'])
        
        # 2. Second table: Version details with instance counts - show ALL versions (no truncation)
        self.write_version_details_table(ws, versions_data['version_counts'])
    
    def write_software_summary_table(self, ws, software_summary):
        """Write table showing software types with counts."""
        # Add table title
        ws['A8'] = "SEoL Software Distribution"
        ws['A8'].font = ws['A8'].font.copy(bold=True)
        
        # Write headers
        headers = ['Software Type', 'Total Instances', 'Unique Versions']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add borders
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Write data with alternating row colors
        total_instances = 0
        total_unique_versions = 0
        
        for i, (_, row) in enumerate(software_summary.iterrows(), 0):
            row_idx = 10 + i
            
            # Alternate row colors
            fill_color = "E9F1FB" if i % 2 == 0 else "FFFFFF"
            
            # Software type cell
            cell = ws.cell(row=row_idx, column=1, value=row['Software'])
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            # Total instances cell
            instances = row['Total Instances']
            cell = ws.cell(row=row_idx, column=2, value=instances)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
            # Unique versions cell
            unique_vers = row['Unique Versions']
            cell = ws.cell(row=row_idx, column=3, value=unique_vers)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
            total_instances += instances
            total_unique_versions += unique_vers
        
        # Add total row
        total_row = 10 + len(software_summary)
        ws.cell(row=total_row, column=1, value="Total")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=total_instances)
        ws.cell(row=total_row, column=2).font = Font(bold=True)
        ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=total_row, column=3, value=total_unique_versions)
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        ws.cell(row=total_row, column=3).alignment = Alignment(horizontal="center")
        
        # Return next available row
        return total_row + 2
    
    def write_version_details_table(self, ws, version_counts):
        """Write table showing detailed version counts for each software - without Version column."""
        # Get next available row
        next_row = 12 + len(version_counts['Software'].unique())
        
        # Add section title
        ws.cell(row=next_row, column=1, value="SEoL Version Details")
        ws.cell(row=next_row, column=1).font = Font(bold=True)
        
        # Write headers - removed 'Version'
        headers = ['Software Type', 'Instance Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=next_row+1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add borders
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Write data with alternating row colors
        # Keep track of current software for grouping
        current_software = None
        
        # Display ALL version entries without showing the version
        for i, (_, row) in enumerate(version_counts.iterrows(), 0):
            row_idx = next_row + 2 + i
            
            # Determine if this is a new software group
            is_new_software = current_software != row['Software']
            current_software = row['Software']
            
            # Alternate row colors within software groups
            fill_color = "E9F1FB" if i % 2 == 0 else "FFFFFF"
            
            # Software cell - only show name for first row of each software type
            software_cell = ws.cell(row=row_idx, column=1, 
                                  value=row['Software'] if is_new_software else "")
            software_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            # Instance count cell - now column 2 instead of 3
            count_cell = ws.cell(row=row_idx, column=2, value=row['Instance Count'])
            count_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            count_cell.alignment = Alignment(horizontal="center")
        
        # Set column widths - removed column B (Version)
        ws.column_dimensions['A'].width = 30  # Software Type
        ws.column_dimensions['B'].width = 15  # Instance Count
        
        # Return the next available row
        return next_row + 2 + len(version_counts)
    
    def create_charts(self, ws, versions_data):
        """Create charts for software and version distribution."""
        # Create software distribution chart
        self.create_software_chart(ws, versions_data['software_counts'])
        
        # Create version distribution chart for top software type
        top_software = versions_data['software_counts'].iloc[0]['Software'] if not versions_data['software_counts'].empty else None
        if top_software:
            software_versions = versions_data['version_counts'][versions_data['version_counts']['Software'] == top_software]
            if not software_versions.empty:
                self.create_version_chart(ws, software_versions, top_software)
    
    def create_software_chart(self, ws, software_counts):
        """Create chart showing software distribution."""
        # Take top 10 software types
        top_software = software_counts.head(10) if len(software_counts) > 10 else software_counts
        
        # Write data for chart
        chart_data_row = 30
        ws.cell(row=chart_data_row, column=7, value="Software")
        ws.cell(row=chart_data_row, column=8, value="Instances")
        
        for i, (_, row) in enumerate(top_software.iterrows(), 1):
            ws.cell(row=chart_data_row+i, column=7, value=row['Software'])
            ws.cell(row=chart_data_row+i, column=8, value=row['Total Instances'])
        
        # Create chart
        chart = BarChart()
        chart.title = "Top SEoL Software by Instance Count"
        chart.y_axis.title = "Number of Instances"
        chart.x_axis.title = "Software Type"
        
        # Set data ranges
        data = Reference(ws, min_col=8, min_row=chart_data_row, max_row=chart_data_row+len(top_software))
        cats = Reference(ws, min_col=7, min_row=chart_data_row+1, max_row=chart_data_row+len(top_software))
        
        # Add data to chart
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Add data labels
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        
        # Increase chart size to avoid overlapping
        chart.width = 25  # Increased from 15
        chart.height = 14  # Increased from 10
        
        # Add chart to worksheet - position at the top right
        ws.add_chart(chart, "D3")
    
    def create_version_chart(self, ws, version_data, software_name):
        """Create chart showing version distribution for top software type."""
        # Take top 8 versions
        top_versions = version_data.head(8) if len(version_data) > 8 else version_data
        
        # Write data for chart
        chart_data_row = 45
        ws.cell(row=chart_data_row, column=7, value="Version")
        ws.cell(row=chart_data_row, column=8, value="Instances")
        
        for i, (_, row) in enumerate(top_versions.iterrows(), 1):
            ws.cell(row=chart_data_row+i, column=7, value=row['Version'])
            ws.cell(row=chart_data_row+i, column=8, value=row['Instance Count'])
        
        # Create chart
        chart = BarChart()
        chart.title = f"{software_name} Version Distribution"
        chart.y_axis.title = "Number of Instances"
        chart.x_axis.title = "Version"
        
        # Set data ranges
        data = Reference(ws, min_col=8, min_row=chart_data_row, max_row=chart_data_row+len(top_versions))
        cats = Reference(ws, min_col=7, min_row=chart_data_row+1, max_row=chart_data_row+len(top_versions))
        
        # Add data to chart
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Add data labels
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        
        # Increase chart size to avoid overlapping
        chart.width = 25  # Increased from 15
        chart.height = 14  # Increased from 10
        
        # Add chart to worksheet - position below the first chart with more space
        ws.add_chart(chart, "D20")